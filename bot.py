
import xml.etree.ElementTree as ET
import base64
import contextlib
import json
import urllib.request
import urllib.parse
import io
import os
import subprocess
import sys
import time
import csv
import re
import threading
import queue
import random
import traceback
import concurrent.futures
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from urllib.parse import urlparse
from collections import Counter

import bruteforce
import fast_dns

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    exit(1)

try:
    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except ImportError:
    print("Установите requests: pip install requests")
    exit(1)

try:
    from yandex_ai_studio_sdk import AIStudio
except ImportError:
    print("Установите SDK: pip install yandex-ai-studio-sdk")
    exit(1)

try:
    from patchright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
    PLAYWRIGHT_OK = True
except ImportError:
    PLAYWRIGHT_OK = False

try:
    from curl_cffi import requests as curl_requests
    CURL_CFFI_OK = True
except ImportError:
    curl_requests = None
    CURL_CFFI_OK = False


def _load_yandex_creds():
    fid = os.environ.get("YANDEX_FOLDER_ID", "").strip()
    key = os.environ.get("YANDEX_API_KEY", "").strip()
    if fid and key:
        return fid, key
    base = os.path.dirname(os.path.abspath(
        sys.executable if getattr(sys, "frozen", False) else __file__))
    try:
        with open(os.path.join(base, "yandex_api.json"), encoding="utf-8") as f:
            d = json.load(f)
        return fid or str(d.get("folder_id", "")).strip(), key or str(d.get("api_key", "")).strip()
    except (OSError, ValueError):
        return fid, key


YANDEX_FOLDER_ID, YANDEX_API_KEY = _load_yandex_creds()

PAGES_PER_QUERY   = 20
RECURSIVE_PAGES   = 25
SEARCH_START_PAGE = 5    # пропускаем первые N страниц выдачи — там обычно мусор
API_DELAY         = 1.5
SEARCH_MODE          = "sync"  # "sync" | "async" — режим по умолчанию, переопределяется из GUI
ASYNC_SUBMIT_DELAY   = 0.3   # сек. пауза между отправкой отложенных запросов (не ждём ответ)
ASYNC_POLL_INTERVAL  = 5     # сек. с каким шагом опрашивать готовность отложенного запроса
ASYNC_OPERATION_TIMEOUT = 10  # сек. на одну отложенную страницу
ASYNC_SEARCH_WORKERS = 10    # сколько отложенных запросов опрашиваем параллельно
ASYNC_SEARCH_START_PAGE = 0
ASYNC_SEARCH_PAGE_COUNT = 13

def _app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


APP_DIR = _app_dir()

ASYNC_QUERIES_DIR = os.path.join(APP_DIR, "async_queries")
ALIVE_TIMEOUT     = 12  # сек. Поднят с 6 — прокси добавляют задержку, и на 6с многие
BRUTEFORCE_ALIVE_TIMEOUT = 3  # сек. В переборе большинство доменов вообще не
MAX_RECURSIVE_DEPTH = 5
ALIVE_WORKERS     = 150  # сколько сайтов проверяем на живость одновременно
BRUTEFORCE_ALIVE_WORKERS = 200
BRUTEFORCE_DISPATCH_WORKERS = 1000
BF_DNS_FLUSH_INTERVAL_SEC = 5  # сек. между авточистками DNS-кеша.
BF_LOG_FLUSH_MS = 150   # мс между отрисовками накопленных строк лога перебора в UI
BF_SPEED_UPDATE_MS = 1000  # мс между пересчётами скорости перебора (доменов/сек) в UI
BF_LOG_MAX_LINES = 5000  # сколько последних строк держим в bf_log_box (защита от разрастания виджета на долгом переборе)
BF_BATCH_SIZE = 1000  # строк за одно чтение из файла пула (захардкожено — без настройки в GUI, см. BruteforceCollector.run())
PLAYWRIGHT_WORKERS = 12  # сколько Chromium-браузеров открываем одновременно для поиска зеркала.
SKIP_PLAYWRIGHT_IF_MIRROR_CONFIRMED = True
MIRROR_CLICK_RETRIES = 3  # доп. попытки клика, если итоговый домен не похож на бренд.
PLAYWRIGHT_HARD_TIMEOUT = 45   # сек. — жёсткий предел на весь вызов Playwright (goto+close),
PLAYWRIGHT_DEADLINE_BUFFER = 17  # сек. запас перед PLAYWRIGHT_HARD_TIMEOUT. Функции
OUTPUT_DIR = "results"

USE_PROXY = False
PROXY_LIST: list = []
PROXY_CONFIG_FILE = os.path.join(
    APP_DIR, "proxy_config.json")


def load_proxy_config():
    global USE_PROXY, PROXY_LIST
    try:
        with open(PROXY_CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        USE_PROXY = bool(cfg.get("use_proxy", USE_PROXY))
        if "proxies" in cfg:
            PROXY_LIST = []
            for p in cfg.get("proxies", []):
                url = str(p.get("url", "")).strip()
                if not url:
                    continue
                entry = {"url": url, "enabled": bool(p.get("enabled", True))}
                if p.get("max_concurrent"):
                    try:
                        entry["max_concurrent"] = int(p["max_concurrent"])
                    except (TypeError, ValueError):
                        pass
                PROXY_LIST.append(entry)
        elif cfg.get("proxy_url"):
            tokens = re.split(r'[\s,;]+', str(cfg["proxy_url"]).strip())
            PROXY_LIST = [{"url": t, "enabled": True} for t in tokens if t]
    except FileNotFoundError:
        pass
    except Exception:
        pass


def save_proxy_config(use_proxy: bool, proxy_list: list):
    try:
        with open(PROXY_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump({"use_proxy": bool(use_proxy), "proxies": proxy_list},
                      f, ensure_ascii=False, indent=2)
    except Exception:
        pass


PLAYWRIGHT_BLOCK_MEDIA = True
_MEDIA_BLOCK_TYPES = {"media", "font"}

IMAGE_REQUIRED_BRANDS = {"Pinco"}
_BLOCKED_URL_HINTS = (
    "google-analytics.com", "googletagmanager.com", "googlesyndication.com",
    "doubleclick.net", "adservice.google", "google.com/ads",
    "mc.yandex.ru", "mc.yandex.com", "yandex.ru/metrika", "an.yandex.ru",
    "top-fwz1.mail.ru", "vk.com/rtrg", "connect.facebook", "facebook.com/tr",
    "facebook.net", "criteo", "hotjar", "clarity.ms", "matomo", "sentry.io",
)


def _looks_like_hostport(s: str) -> bool:
    host, sep, port = s.rpartition(":")
    return bool(sep) and port.isdigit() and "." in host


def _normalize_proxy(raw: str) -> str:
    raw = raw.strip()
    if not raw:
        return ""
    scheme = "http"
    if "://" in raw:
        scheme, raw = raw.split("://", 1)
    if "@" in raw:
        left, right = raw.rsplit("@", 1)
        if _looks_like_hostport(left) and not _looks_like_hostport(right):
            hostport, creds = left, right
        else:
            creds, hostport = left, right
        return f"{scheme}://{creds}@{hostport}" if creds else f"{scheme}://{hostport}"
    parts = raw.split(":")
    if len(parts) == 4:  # ip:port:user:pass
        host, port, user, pw = parts
        return f"{scheme}://{user}:{pw}@{host}:{port}"
    return f"{scheme}://{raw}"


def _proxy_pool() -> list:
    if not USE_PROXY:
        return []
    return [n for n in (_normalize_proxy(p["url"]) for p in PROXY_LIST
                        if p.get("enabled")) if n]


BLOCK_STATUS_CODES = {403, 429}
MAX_PROXY_RETRIES = 4  # сколько разных прокси пробуем на одну ссылку, прежде чем сдаться


_proxy_rr_lock = threading.Lock()
_proxy_rr_counter = 0

_proxy_pin = threading.local()


@contextlib.contextmanager
def _pinned_proxy(proxy_url: str):
    prev = getattr(_proxy_pin, "value", None)
    _proxy_pin.value = proxy_url
    try:
        yield
    finally:
        _proxy_pin.value = prev


def _pick_proxy(exclude: set = None) -> str:
    global _proxy_rr_counter
    pinned = getattr(_proxy_pin, "value", None)
    if pinned and not (exclude and pinned in exclude):
        return pinned
    pool = _proxy_pool()
    if exclude:
        pool = [p for p in pool if p not in exclude]
    if not pool:
        return ""
    healthy = [p for p in pool if _is_proxy_healthy(p)]
    if healthy:
        pool = healthy
    with _proxy_rr_lock:
        idx = _proxy_rr_counter % len(pool)
        _proxy_rr_counter += 1
    return pool[idx]


def _proxy_label(proxy_url: str) -> str:
    p = urlparse(proxy_url)
    return f"{p.hostname}:{p.port}" if p.port else (p.hostname or proxy_url)


def _requests_proxies(exclude: set = None):
    p = _pick_proxy(exclude)
    return {"http": p, "https": p} if p else None


def _playwright_proxy_dict(proxy_url: str):
    if not proxy_url:
        return None
    p = urlparse(proxy_url)
    if not p.hostname:
        return None
    server = f"{p.scheme}://{p.hostname}" + (f":{p.port}" if p.port else "")
    proxy = {"server": server}
    if p.username:
        proxy["username"] = urllib.parse.unquote(p.username)
    if p.password:
        proxy["password"] = urllib.parse.unquote(p.password)
    return proxy


def _playwright_proxy(exclude: set = None):
    return _playwright_proxy_dict(_pick_proxy(exclude))


def _apply_media_block(context, brand: str = "", url: str = ""):
    if not PLAYWRIGHT_BLOCK_MEDIA:
        return
    blocked_types = set(_MEDIA_BLOCK_TYPES)
    if _effective_brand(brand, url) not in IMAGE_REQUIRED_BRANDS:
        blocked_types.add("image")

    def handler(route):
        req = route.request
        if (req.resource_type in blocked_types
                or any(h in req.url for h in _BLOCKED_URL_HINTS)):
            route.abort()
        else:
            route.continue_()
    try:
        context.route("**/*", handler)
    except Exception:
        pass


def _requests_get_retry_proxy(url: str, headers: dict, timeout: int,
                               allow_health_wait: bool = True,
                               mobile=None):
    mobile = _use_mobile(mobile)
    pool = _proxy_pool()
    tried: set = set()
    max_tries = max(1, min(MAX_PROXY_RETRIES, len(pool))) if pool else 1
    resp = None
    last_exc = None
    for _ in range(max_tries):
        proxy = _pick_proxy(exclude=tried)
        if pool:
            tried.add(proxy)
        proxies = {"http": proxy, "https": proxy} if proxy else None
        profiles = MOBILE_CURL_CFFI_PROFILES if mobile else CURL_CFFI_PROFILES
        try:
            if CURL_CFFI_OK and profiles:
                profile = random.choice(list(profiles))
                if mobile:
                    req_headers = _apply_mobile_ua(headers, profiles[profile])
                else:
                    req_headers = dict(headers)
                    req_headers["User-Agent"] = profiles[profile]
                resp = curl_requests.get(
                    url, timeout=timeout, headers=req_headers,
                    allow_redirects=True, verify=False, proxies=proxies,
                    impersonate=profile,
                )
            else:
                resp = requests.get(
                    url, timeout=timeout, headers=headers,
                    allow_redirects=True, verify=False, proxies=proxies,
                )
        except Exception as e:
            last_exc = e
            if not pool:
                raise last_exc
            continue
        if resp.status_code not in BLOCK_STATUS_CODES:
            return resp
        if not pool or len(tried) >= len(pool):
            return resp
    if resp is not None:
        return resp
    if (allow_health_wait and pool and last_exc is not None
            and _all_unhealthy(pool) and _wait_for_any_healthy_proxy(pool)):
        return _requests_get_retry_proxy(url, headers, timeout,
                                         allow_health_wait=False, mobile=mobile)
    raise last_exc


def _requests_get_via_search(url: str, referer_base: str, seed: str,
                              ua: str, timeout: int, allow_health_wait: bool = True,
                              mobile=None):
    mobile = _use_mobile(mobile)
    referer_url = _realistic_referer(referer_base, seed)
    pool = _proxy_pool()
    tried: set = set()
    max_tries = max(1, min(MAX_PROXY_RETRIES, len(pool))) if pool else 1
    resp = None
    last_exc = None
    for _ in range(max_tries):
        proxy = _pick_proxy(exclude=tried)
        if pool:
            tried.add(proxy)
        proxies = {"http": proxy, "https": proxy} if proxy else None
        target_headers = _build_headers(
            referer=referer_url,
            user_agent=None if (CURL_CFFI_OK or mobile) else ua,
            mobile=mobile)
        profiles = MOBILE_CURL_CFFI_PROFILES if mobile else CURL_CFFI_PROFILES
        try:
            if CURL_CFFI_OK and profiles:
                profile = random.choice(list(profiles))
                if mobile:
                    target_headers = _apply_mobile_ua(target_headers, profiles[profile])
                else:
                    target_headers["User-Agent"] = profiles[profile]
                resp = curl_requests.get(
                    url, timeout=timeout, headers=target_headers,
                    allow_redirects=True, verify=False, proxies=proxies,
                    impersonate=profile,
                )
            else:
                resp = requests.get(
                    url, timeout=timeout, headers=target_headers,
                    allow_redirects=True, verify=False, proxies=proxies,
                )
        except Exception as e:
            last_exc = e
            if not pool:
                raise last_exc
            continue
        if resp.status_code not in BLOCK_STATUS_CODES:
            return resp, referer_url
        if not pool or len(tried) >= len(pool):
            return resp, referer_url
    if resp is not None:
        return resp, referer_url
    if (allow_health_wait and pool and last_exc is not None
            and _all_unhealthy(pool) and _wait_for_any_healthy_proxy(pool)):
        return _requests_get_via_search(url, referer_base, seed, ua, timeout,
                                         allow_health_wait=False, mobile=mobile)
    raise last_exc


def _get_via_proxy_for_test(url: str, proxy: str, timeout: int):
    proxies = {"http": proxy, "https": proxy}
    if CURL_CFFI_OK:
        profile = random.choice(list(CURL_CFFI_PROFILES))
        return curl_requests.get(
            url, timeout=timeout,
            headers={"User-Agent": CURL_CFFI_PROFILES[profile]},
            verify=False, proxies=proxies, impersonate=profile)
    return requests.get(
        url, timeout=timeout, proxies=proxies, verify=False,
        headers={"User-Agent": "Mozilla/5.0"})


PROXY_HEALTH_CHECK_URL = "https://api.ipify.org?format=json"
PROXY_HEALTH_CHECK_INTERVAL = 5     # сек. между хендшейками одного прокси
PROXY_HEALTH_MAX_WAIT = 180         # сек. максимум ждём восстановления прокси
PROXY_HEALTH_POLL_INTERVAL = 2      # сек. между проверками "уже здоров?"

_proxy_healthy: dict = {}
_proxy_health_lock = threading.Lock()
_proxy_health_monitor_started = False
_proxy_health_monitor_lock = threading.Lock()

_stop_event = threading.Event()


def _probe_proxy_once(proxy: str, timeout: float = 8) -> bool:
    try:
        _get_via_proxy_for_test(PROXY_HEALTH_CHECK_URL, proxy, timeout)
        return True
    except Exception:
        return False


def _proxy_health_monitor_loop(log_fn):
    while True:
        for proxy in _proxy_pool():
            ok = _probe_proxy_once(proxy)
            with _proxy_health_lock:
                prev = _proxy_healthy.get(proxy)
                _proxy_healthy[proxy] = ok
            if prev is not None and prev != ok:
                label = _proxy_label(proxy)
                if ok:
                    log_fn(f"[прокси] {label}: снова отвечает (похоже, сменился IP)")
                else:
                    log_fn(f"[прокси] {label}: не отвечает (похоже на смену IP) — "
                           f"проверки через него временно приостанавливаются")
        time.sleep(PROXY_HEALTH_CHECK_INTERVAL)


def start_proxy_health_monitor(log_fn=print):
    global _proxy_health_monitor_started
    with _proxy_health_monitor_lock:
        if _proxy_health_monitor_started:
            return
        _proxy_health_monitor_started = True
    threading.Thread(target=_proxy_health_monitor_loop, args=(log_fn,),
                      daemon=True, name="proxy-health").start()


def _is_proxy_healthy(proxy: str) -> bool:
    with _proxy_health_lock:
        return _proxy_healthy.get(proxy, True)


def _all_unhealthy(pool: list) -> bool:
    return bool(pool) and all(not _is_proxy_healthy(p) for p in pool)


def _wait_for_any_healthy_proxy(pool: list, log_fn=print) -> bool:
    deadline = time.time() + PROXY_HEALTH_MAX_WAIT
    logged = False
    while time.time() < deadline:
        if _stop_event.is_set():
            return False
        if not _all_unhealthy(pool):
            return True
        if not logged:
            log_fn("[прокси] все прокси сейчас не отвечают (похоже на смену IP) — "
                   "жду восстановления, проверки на паузе...")
            logged = True
        time.sleep(PROXY_HEALTH_POLL_INTERVAL)
    return not _all_unhealthy(pool)


def test_proxy(log_fn=print) -> bool:
    pool = _proxy_pool()
    if not pool:
        return True
    log_fn(f"[*] Проверяю прокси ({len(pool)} шт.)...")
    ok = 0
    for pr in pool:
        try:
            resp = _get_via_proxy_for_test("https://api.ipify.org?format=json", pr, 15)
            ip = resp.json().get("ip", "?")
            log_fn(f"[+] {_proxy_label(pr)} → внешний IP {ip}")
            ok += 1
        except Exception as e:
            log_fn(f"[!] {_proxy_label(pr)} не работает: {e}")
    log_fn(f"[*] Рабочих прокси: {ok}/{len(pool)}")
    return ok > 0


def test_single_proxy(raw_url: str, timeout: int = 12) -> tuple:
    normalized = _normalize_proxy(raw_url)
    if not normalized:
        return False, "пустой адрес"
    try:
        resp = _get_via_proxy_for_test("https://api.ipify.org?format=json", normalized, timeout)
        return True, resp.json().get("ip", "?")
    except Exception as e:
        return False, str(e)


_playwright_executor = concurrent.futures.ThreadPoolExecutor(
    max_workers=PLAYWRIGHT_WORKERS, thread_name_prefix="pw")

GOOGLE_SHEET_CSV = (
    "https://docs.google.com/spreadsheets/d/"
    "1GihafERPUTyiI-X_b9VEkNTNbplwHLdd"
    "/export?format=csv&gid=1098631008"
)
GOOGLE_SHEET_DOMAIN_COL = 1  # столбец B (без заголовка)

TRASH_SHEET_CSV = (
    "https://docs.google.com/spreadsheets/d/"
    "1e2nacNzVm6W3IA3EwTESAjYd5q5JQdVqwMi4TZMAv9k"
    "/export?format=csv&gid=0"
)

RESULTS_APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbwu-h7A9VXPcO76G8wiFRjD42ZHJivdtUingg02H-pSLQNrGnA6lCPup5ceNd8B05mw/exec"


BRAND_STRICT_PREFIXES = {
    "1xBet":     ["1xbet", "1xlite"],
    "Melbet":    ["melbet"],
    "CatCasino": ["catcasino", "cat-casino", "cat-kazino"],
    "Mostbet":   ["mostbet"],
    "Pinco":     ["pinco", "pin-up", "pinup"],
    "Vavada":    ["vavada"],
    "1xCasino":  ["1xcasino"],
}

BRAND_QUERIES = {
    "1xBet": [
        "https://1xbet-app-bet.ru/",
        "https://1xbet-sisk.top/",
        "https://1xbet-nyno.top/",
        "https://1xbet-t5kv.cfd/",
        "https://1xbet-s1k7e.xyz/",
        "https://1xbet-o0s06.cyou/",
        "1xbet рабочий сайт",
        "1xbet альтернативный адрес",
        "1xbet обход блокировки",
        "1xbet зеркало актуальное",
    ],
    "Melbet": [
        "melbet зеркало",
        "melbet зеркало рабочее",
        "melbet зеркало на сегодня",
        "melbet рабочее зеркало прямо сейчас",
        "melbet официальный сайт зеркало",
        "melbet вход через зеркало",
        "melbet рабочий сайт",
        "melbet альтернативный адрес",
        "melbet обход блокировки",
        "melbet зеркало актуальное",
    ],
    "CatCasino": [
        "cat casino зеркало",
        "catcasino зеркало рабочее",
        "cat casino зеркало на сегодня",
        "catcasino рабочее зеркало",
        "cat casino официальный сайт",
        "cat casino вход",
        "catcasino рабочий сайт",
        "cat casino альтернативный адрес",
        "cat casino обход блокировки",
        "catcasino зеркало актуальное",
    ],
    "Mostbet": [
        "mostbet зеркало",
        "mostbet зеркало рабочее",
        "mostbet зеркало на сегодня",
        "mostbet рабочее зеркало прямо сейчас",
        "mostbet официальный сайт зеркало",
        "mostbet вход через зеркало",
        "mostbet рабочий сайт",
        "mostbet альтернативный адрес",
        "mostbet обход блокировки",
        "mostbet зеркало актуальное",
    ],
    "Pinco": [
        "pinco зеркало",
        "pinco зеркало рабочее",
        "pin-up зеркало на сегодня",
        "pinup рабочее зеркало",
        "pinco официальный сайт",
        "pin-up вход через зеркало",
        "pinco рабочий сайт",
        "pinup альтернативный адрес",
        "pin-up обход блокировки",
        "pinco зеркало актуальное",
    ],
    "Vavada": [
        "vavada зеркало",
        "vavada зеркало рабочее",
        "vavada зеркало на сегодня",
        "vavada рабочее зеркало прямо сейчас",
        "vavada официальный сайт зеркало",
        "vavada вход через зеркало",
        "vavada рабочий сайт",
        "vavada альтернативный адрес",
        "vavada обход блокировки",
        "vavada зеркало актуальное",
    ],
    "1xCasino": [
        "1xcasino зеркало",
        "1xcasino зеркало рабочее",
        "1xcasino зеркало на сегодня",
        "1xcasino рабочее зеркало прямо сейчас",
        "1xcasino официальный сайт зеркало",
        "1xcasino вход через зеркало",
        "1xcasino рабочий сайт",
        "1xcasino альтернативный адрес",
        "1xcasino обход блокировки",
        "1xcasino зеркало актуальное",
    ],
}

GENERAL_BRAND = "Общий"
BRAND_QUERIES[GENERAL_BRAND] = [q for brand in BRAND_STRICT_PREFIXES for q in BRAND_QUERIES[brand]]

BRAND_NAMES = list(BRAND_STRICT_PREFIXES.keys()) + [GENERAL_BRAND]
FILTER_LEVELS = ["Лёгкий", "Строгий"]


def load_async_queries(brand: str, log_fn=print) -> list[str]:
    path = os.path.join(ASYNC_QUERIES_DIR, f"{brand}.txt")
    if not os.path.exists(path):
        log_fn(f"[!] Файл запросов не найден: {path}")
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            lines = f.readlines()
    except Exception as e:
        log_fn(f"[!] Ошибка чтения файла запросов {path}: {e}")
        return []
    queries = [ln.strip() for ln in lines]
    queries = [q for q in queries if q and not q.startswith("#")]
    log_fn(f"[*] Загружено {len(queries)} запросов из {os.path.basename(path)}")
    return queries


def extract_domain(url: str) -> str:
    try:
        h = urlparse(url).hostname or ""
        return h.removeprefix("www.")
    except Exception:
        return ""


def _parse_domain_cell(val: str) -> str:
    val = val.strip().lower()
    if not val:
        return ""
    if "://" not in val:
        val = "https://" + val
    try:
        h = urlparse(val).hostname or ""
        return h.removeprefix("www.")
    except Exception:
        return val


def has_repeated_locale(url: str, threshold: int = 5) -> bool:
    return bool(re.search(r'(/[a-z]{2})\1{' + str(threshold - 1) + r',}', url, re.IGNORECASE))


def _has_anagram_window(text: str, word: str) -> bool:
    n, m = len(text), len(word)
    if m > n:
        return False
    target = Counter(word)
    window = Counter(text[:m])
    if window == target:
        return True
    for i in range(m, n):
        window[text[i]] += 1
        drop = text[i - m]
        window[drop] -= 1
        if window[drop] == 0:
            del window[drop]
        if window == target:
            return True
    return False


def _brand_match(url: str, directs: list, anagram_word: str = None,
                 combined=None) -> bool:
    u = url.lower()
    for kw in directs:
        if kw in u:
            return True
    if combined:
        prefix, suffixes = combined
        if prefix in u and any(s in u for s in suffixes):
            return True
    if anagram_word:
        clean = "".join(c for c in u if c.isalpha())
        if _has_anagram_window(clean, anagram_word):
            return True
    return False


BRAND_MATCHERS = {
    "1xBet": lambda url: _brand_match(url, ["1xbet", "1xlite"]),
    "Melbet": lambda url: _brand_match(url, ["melbet"], "melbet"),
    "CatCasino": lambda url: _brand_match(
        url, ["catcasino", "cat-casino", "cat-kazino"],
        "catcasino", combined=("cat", ["casino", "kazino"])),
    "Mostbet": lambda url: _brand_match(url, ["mostbet"], "mostbet"),
    "Pinco": lambda url: _brand_match(url, ["pinco", "pinup", "pin-up"], "pinco"),
    "Vavada": lambda url: _brand_match(url, ["vavada"], "vavada"),
    "1xCasino": lambda url: _brand_match(url, ["1xcasino"]),
}


def _strict_brand_match(url: str, prefixes: list) -> bool:
    try:
        host = urlparse(url).hostname or ""
    except Exception:
        host = ""
    host = host.lower().removeprefix("www.")
    return any(host.startswith(p) for p in prefixes)


def matches_brand(url: str, brand: str, strict: bool = True) -> bool:
    if strict:
        prefixes = BRAND_STRICT_PREFIXES.get(brand)
        if prefixes is None:
            return True
        return _strict_brand_match(url, prefixes)
    else:
        matcher = BRAND_MATCHERS.get(brand)
        if matcher is None:
            return True
        return matcher(url)


def _match_any_brand(url: str, strict: bool) -> str:
    for name in BRAND_STRICT_PREFIXES:
        if matches_brand(url, name, strict=strict):
            return name
    return ""


TRACKER_DOMAINS = {
    "call2me.pro",
    "sendmad.org",
    "traffic-boost-offer.com",
    "linksyraflow.com",
    "adsbestuse.com",
    "alph-2apex.com",
    "apx3-smart.com",
    "fastslotus-route.com",
    "gglinks.top",
    "onxtrafficjump.com",
    "slotbonususe.com",
    "routecoreuse.com",
    "stepintowinnerzone.com",
    "i01yard.win",
    "rbt-itk",
    "on-x",
    "normcasino",
    "iris-vel0ra2.com",
    'pegasus-velor4-3.com',
    'safar1-4velora.com',
    'gr4ce-v3l0ra.com',
}

TRACKER_DOMAIN_PATTERNS = [
    re.compile(r'^7k\d+\.casino$', re.IGNORECASE),
]


def is_tracker_domain(url: str) -> bool:
    host = extract_domain(url)
    if not host:
        return False
    if any(host == d or host.endswith("." + d) for d in TRACKER_DOMAINS):
        return True
    return any(p.match(host) for p in TRACKER_DOMAIN_PATTERNS)


AUTO_APPROVE_HOST_PATTERNS = [
    re.compile(r'^pin-up-casino-[a-z0-9]+\.top$', re.IGNORECASE),
]


def is_auto_approved(url: str) -> bool:
    if not url:
        return False
    host = extract_domain(url if "://" in url else "https://" + url)
    if not host:
        return False
    return any(p.match(host) for p in AUTO_APPROVE_HOST_PATTERNS)


DEFAULT_REQUIRED_MIRROR_KEYWORDS = (
    "1xlite", "melbet", "pinup", "pinco",
    "adavav", "mostbet",
    "catcasino", "cat-casino", "cat-kazino", "trovina", "vaaaavaaada", "acz6uomb", "zwjfsgmost", "vaaaavaaadaq", "krv8g2nkvv", "vadava2ja6y", "trovina.dev",
    "1xcasino","vadavasj9","vadava",'esnc9bsxmb',"vcjwmr35mb","weg96sbmb","kcxw71rsmb","trovina","29upq21nmb.com",'kcxw71rsmb.com',"pincte.com"
)

REQUIRED_MIRROR_KEYWORDS = list(DEFAULT_REQUIRED_MIRROR_KEYWORDS)

MIRROR_KEYWORDS_FILE = os.path.join(
    APP_DIR, "mirror_keywords.json")


def _clean_mirror_keywords(raw) -> list:
    out, seen = [], set()
    for item in raw or []:
        kw = str(item).strip().lower()
        if not kw or kw in seen:
            continue
        seen.add(kw)
        out.append(kw)
    return out


def load_mirror_keywords():
    global REQUIRED_MIRROR_KEYWORDS
    try:
        with open(MIRROR_KEYWORDS_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except FileNotFoundError:
        return
    except Exception:
        return
    if isinstance(cfg, dict):
        cfg = cfg.get("keywords")
    if isinstance(cfg, list):
        REQUIRED_MIRROR_KEYWORDS = _clean_mirror_keywords(cfg)


def save_mirror_keywords(keywords: list) -> bool:
    try:
        with open(MIRROR_KEYWORDS_FILE, "w", encoding="utf-8") as f:
            json.dump({"keywords": list(keywords)}, f,
                      ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def set_mirror_keywords(keywords: list) -> tuple:
    global REQUIRED_MIRROR_KEYWORDS
    REQUIRED_MIRROR_KEYWORDS = _clean_mirror_keywords(keywords)
    ok = save_mirror_keywords(REQUIRED_MIRROR_KEYWORDS)
    return REQUIRED_MIRROR_KEYWORDS, ok


load_mirror_keywords()


BF_QUEUE_FILE = os.path.join(
    APP_DIR, "bruteforce_queue.json")


def _clean_bf_task(raw) -> dict:
    if not isinstance(raw, dict):
        return {}
    try:
        length = int(raw.get("length", 0))
    except (TypeError, ValueError):
        return {}
    if length < 1:
        return {}
    charset = str(raw.get("charset") or "letters_digits")
    if charset not in bruteforce.CHARSETS:
        charset = "letters_digits"
    brand = str(raw.get("brand") or "")
    if brand not in BRAND_NAMES:
        brand = BRAND_NAMES[0]
    return {"prefix": str(raw.get("prefix") or ""),
            "suffix": str(raw.get("suffix") or ""),
            "length": length, "charset": charset, "brand": brand}


def bf_task_label(task: dict) -> str:
    mask = f"{task['prefix']}{'X' * task['length']}{task['suffix']}"
    return (f"{mask}  |  {bruteforce.CHARSET_LABELS.get(task['charset'], task['charset'])}"
            f"  |  {task['brand']}")


def load_bf_queue() -> list:
    try:
        with open(BF_QUEUE_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except FileNotFoundError:
        return []
    except Exception:
        return []
    if isinstance(cfg, dict):
        cfg = cfg.get("tasks")
    if not isinstance(cfg, list):
        return []
    return [t for t in (_clean_bf_task(x) for x in cfg) if t]


def save_bf_queue(tasks: list) -> bool:
    try:
        with open(BF_QUEUE_FILE, "w", encoding="utf-8") as f:
            json.dump({"tasks": list(tasks)}, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def _looks_like_real_mirror(url: str) -> bool:
    return bool(
        url
        and not is_tracker_domain(url)
        and not has_repeated_locale(url)
        and _mirror_has_required_keyword(url)
    )


def _row_auto_approved(r: dict) -> bool:
    return (is_auto_approved(r.get("url", ""))
            and r.get("alive_info", {}).get("error") != "dns_not_found")


def _auto_approved_alive(url: str, alive_info: dict) -> bool:
    return bool(
        is_auto_approved(url)
        and alive_info.get("alive")
        and alive_info.get("status_code") == 200
    )


def _is_mirror_confirmed(mirror: str, final: str, brand: str) -> bool:
    return bool(
        SKIP_PLAYWRIGHT_IF_MIRROR_CONFIRMED
        and mirror and mirror != final
        and _looks_like_real_mirror(mirror)
        and (matches_brand(mirror, brand, strict=False)
             or _match_any_brand(mirror, strict=False))
    )


def _mirror_has_required_keyword(url: str) -> bool:
    if not url:
        return False
    if is_auto_approved(url):
        return True
    u = url.lower()
    return any(kw in u for kw in REQUIRED_MIRROR_KEYWORDS)


def load_known_domains(log_fn=print) -> set:
    log_fn("[*] Загружаю базу доменов из Google Sheets...")
    try:
        req = urllib.request.Request(
            GOOGLE_SHEET_CSV, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            text = resp.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        next(reader, None)  # заголовок (колонка B без названия)
        col_idx = GOOGLE_SHEET_DOMAIN_COL
        domains = set()
        for row in reader:
            val = row[col_idx] if col_idx < len(row) else None
            if val:
                d = _parse_domain_cell(str(val).strip())
                if d:
                    domains.add(d)
        log_fn(f"[+] Загружено {len(domains)} известных доменов")
        return domains
    except Exception as e:
        log_fn(f"[!] Ошибка загрузки Google Sheets: {e}")
        return set()


def load_trash_domains(log_fn=print) -> set:
    log_fn("[*] Загружаю мусорную базу...")
    try:
        req = urllib.request.Request(
            TRASH_SHEET_CSV, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            text = resp.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        header = next(reader, None)
        if not header:
            return set()
        col_idx = next(
            (i for i, c in enumerate(header)
             if c and c.strip().lower() in ("домены", "домен", "domain")), None)
        if col_idx is None:
            col_idx = 1
        domains = set()
        for row in reader:
            if col_idx < len(row):
                d = _parse_domain_cell(row[col_idx])
                if d:
                    domains.add(d)
        log_fn(f"[+] Загружено {len(domains)} мусорных доменов")
        return domains
    except Exception as e:
        log_fn(f"[!] Ошибка загрузки мусорной базы: {e}")
        return set()


def load_results_domains(log_fn=print) -> set:
    if not RESULTS_APPS_SCRIPT_URL:
        return set()
    log_fn("[*] Загружаю таблицу результатов для дедупа...")
    try:
        req = urllib.request.Request(
            RESULTS_APPS_SCRIPT_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = resp.read().decode("utf-8", "replace")
        data = json.loads(body)
        domains = set()
        for u in data.get("urls", []):
            d = _parse_domain_cell(str(u))
            if d:
                domains.add(d)
        log_fn(f"[+] В таблице результатов уже {len(domains)} доменов")
        return domains
    except Exception as e:
        log_fn(f"[!] Ошибка загрузки таблицы результатов: {e}")
        return set()


LOCAL_REJECTED_FILE = os.path.join(
    APP_DIR, "rejected_domains.txt")

_local_rejected_lock = threading.Lock()


def load_local_rejected(log_fn=print) -> set:
    if not os.path.exists(LOCAL_REJECTED_FILE):
        return set()
    try:
        with open(LOCAL_REJECTED_FILE, "r", encoding="utf-8") as f:
            domains = {ln.strip() for ln in f if ln.strip() and not ln.startswith("#")}
        log_fn(f"[+] Локальный список отклонённых: {len(domains)} доменов")
        return domains
    except Exception as e:
        log_fn(f"[!] Ошибка чтения локального списка отклонённых: {e}")
        return set()


def append_local_rejected(domain: str):
    with _local_rejected_lock:
        try:
            with open(LOCAL_REJECTED_FILE, "a", encoding="utf-8") as f:
                f.write(domain + "\n")
        except Exception:
            pass


def _init_sdk():
    return AIStudio(folder_id=YANDEX_FOLDER_ID, auth=YANDEX_API_KEY)


SEARCH_REGIONS = [
    ("225", "Россия"), ("187", "Украина"), ("149", "Беларусь"), ("159", "Казахстан"),
    ("213", "Москва"), ("2", "Санкт-Петербург"), ("1", "Москва и область"),
    ("20", "Архангельск"), ("37", "Астрахань"), ("197", "Барнаул"), ("4", "Белгород"),
    ("77", "Благовещенск"), ("191", "Брянск"), ("24", "Великий Новгород"),
    ("75", "Владивосток"), ("33", "Владикавказ"), ("192", "Владимир"),
    ("38", "Волгоград"), ("21", "Вологда"), ("193", "Воронеж"), ("1106", "Грозный"),
    ("54", "Екатеринбург"), ("5", "Иваново"), ("63", "Иркутск"), ("41", "Йошкар-Ола"),
    ("43", "Казань"), ("22", "Калининград"), ("64", "Кемерово"), ("7", "Кострома"),
    ("35", "Краснодар"), ("62", "Красноярск"), ("53", "Курган"), ("8", "Курск"),
    ("9", "Липецк"), ("28", "Махачкала"), ("1092", "Назрань"), ("30", "Нальчик"),
    ("47", "Нижний Новгород"), ("65", "Новосибирск"), ("66", "Омск"), ("10", "Орел"),
    ("48", "Оренбург"), ("49", "Пенза"), ("50", "Пермь"), ("25", "Псков"),
    ("39", "Ростов-на-Дону"), ("11", "Рязань"), ("51", "Самара"), ("42", "Саранск"),
    ("12", "Смоленск"), ("239", "Сочи"), ("36", "Ставрополь"), ("973", "Сургут"),
    ("13", "Тамбов"), ("14", "Тверь"), ("67", "Томск"), ("15", "Тула"),
    ("195", "Ульяновск"), ("172", "Уфа"), ("76", "Хабаровск"), ("45", "Чебоксары"),
    ("56", "Челябинск"), ("1104", "Черкесск"), ("16", "Ярославль"), ("23", "Мурманск"),
]

SEARCH_REGION_MODE = "rotate"
SEARCH_REGION_FIXED_ID = None   # используется только при SEARCH_REGION_MODE == "fixed"

_region_rr_lock = threading.Lock()
_region_rr_counter = 0


def _pick_search_region() -> tuple:
    global _region_rr_counter
    if SEARCH_REGION_MODE == "none" or not SEARCH_REGIONS:
        return None, None
    if SEARCH_REGION_MODE == "fixed":
        if not SEARCH_REGION_FIXED_ID:
            return None, None
        for rid, label in SEARCH_REGIONS:
            if rid == SEARCH_REGION_FIXED_ID:
                return rid, label
        return None, None
    with _region_rr_lock:
        idx = _region_rr_counter % len(SEARCH_REGIONS)
        _region_rr_counter += 1
    return SEARCH_REGIONS[idx]


def _parse_search_xml(xml_data: bytes, log_fn=print) -> list[dict]:
    results = []
    try:
        root = ET.fromstring(xml_data)
        error = root.find(".//error")
        if error is not None:
            log_fn(f"[!] API ошибка: {error.text}")
            return []
        for group in root.findall(".//group"):
            doc = group.find("doc")
            if doc is None:
                continue
            doc_url = doc.findtext("url", "")
            doc_domain = doc.findtext("domain", "")
            doc_title = doc.findtext("title", "")
            passage = doc.findtext(".//passage", "")
            doc_title = re.sub(r'<[^>]+>', '', doc_title)
            passage = re.sub(r'<[^>]+>', '', passage)
            if doc_url:
                results.append({
                    "url": doc_url,
                    "domain": doc_domain or extract_domain(doc_url),
                    "title": doc_title,
                    "snippet": passage,
                })
    except ET.ParseError as e:
        log_fn(f"[!] Ошибка парсинга XML: {e}")
    return results


def yandex_search(sdk, query: str, page: int = 0, log_fn=print, region: str = None) -> list[dict]:
    try:
        kwargs = {"region": region} if region else {}
        search = sdk.search_api.web(search_type="ru", **kwargs)
        raw_xml = search.run(query, format="xml", page=page)
        if isinstance(raw_xml, bytes):
            xml_data = raw_xml
        else:
            xml_data = raw_xml.encode("utf-8")
    except Exception as e:
        log_fn(f"[!] Ошибка API: {e}")
        return []
    return _parse_search_xml(xml_data, log_fn)


def yandex_search_submit(sdk, query: str, page: int = 0, region: str = None):
    kwargs = {"region": region} if region else {}
    search = sdk.search_api.web(search_type="ru", **kwargs)
    return search.run_deferred(query, format="xml", page=page)


def yandex_search_await(operation, log_fn=print, poll_interval: float = 5,
                        poll_timeout: float = ASYNC_OPERATION_TIMEOUT) -> list[dict]:
    try:
        raw_xml = operation.wait(poll_interval=poll_interval, poll_timeout=poll_timeout)
        if isinstance(raw_xml, bytes):
            xml_data = raw_xml
        else:
            xml_data = raw_xml.encode("utf-8")
    except Exception as e:
        log_fn(f"[!] Ошибка отложенного API (или истёк тайм-аут {poll_timeout}с): {e}")
        return []
    return _parse_search_xml(xml_data, log_fn)


ALIVE_REFERERS = [
    "https://yandex.ru/search/",
    "https://www.yandex.ru/search",
    "https://ya.ru/search",
    "https://www.google.com/search",
]

def _realistic_referer(base: str, seed: str) -> str:
    text = (seed or "site").strip()
    if "yandex" in base or "ya.ru" in base:
        q = urllib.parse.urlencode({"text": text, "lr": "213"})
    else:
        q = urllib.parse.urlencode({"q": text})
    sep = "&" if "?" in base else "?"
    return f"{base}{sep}{q}"


USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
]

PLAYWRIGHT_VIEWPORTS = [
    {"width": 1366, "height": 768},
    {"width": 1920, "height": 1080},
    {"width": 1536, "height": 864},
    {"width": 1440, "height": 900},
    {"width": 1280, "height": 800},
]


def _random_user_agent() -> str:
    return random.choice(USER_AGENTS)


def _random_viewport() -> dict:
    return random.choice(PLAYWRIGHT_VIEWPORTS)


CURL_CFFI_PROFILES = {
    "chrome124": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "chrome123": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "chrome120": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "chrome119": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
}


FORCE_MOBILE = True

REQUESTS_DEVICE_FALLBACK = True    # requests-фаза (Фаза 3)
PLAYWRIGHT_DEVICE_FALLBACK = True  # клик в браузере (Фаза 4)


def _use_mobile(mobile) -> bool:
    return FORCE_MOBILE if mobile is None else bool(mobile)

CLOAK_EMPTY_BODY_LIMIT = 200

MOBILE_USER_AGENTS = [
    "Mozilla/5.0 (Linux; Android 14; Pixel 8) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; Android 13; SM-A536B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; Android 14; SM-S918B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 18_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.0 Mobile/15E148 Safari/604.1",
]

MOBILE_DEVICES = [
    {"user_agent": MOBILE_USER_AGENTS[0], "viewport": {"width": 412, "height": 915},
     "device_scale_factor": 2.625, "is_mobile": True, "has_touch": True},
    {"user_agent": MOBILE_USER_AGENTS[1], "viewport": {"width": 360, "height": 800},
     "device_scale_factor": 3, "is_mobile": True, "has_touch": True},
    {"user_agent": MOBILE_USER_AGENTS[2], "viewport": {"width": 393, "height": 852},
     "device_scale_factor": 3, "is_mobile": True, "has_touch": True},
]

_MOBILE_CURL_CFFI_PROFILES_ALL = {
    "chrome131_android": "Mozilla/5.0 (Linux; Android 10; K) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Mobile Safari/537.36",
    "safari180_ios": "Mozilla/5.0 (iPhone; CPU iPhone OS 18_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.0 Mobile/15E148 Safari/604.1",
}


def _supported_curl_cffi_profiles(profiles: dict) -> dict:
    if not CURL_CFFI_OK:
        return {}
    try:
        from curl_cffi.requests import BrowserType
        available = {b.value for b in BrowserType}
    except Exception:
        return dict(profiles)
    return {k: v for k, v in profiles.items() if k in available}


MOBILE_CURL_CFFI_PROFILES = _supported_curl_cffi_profiles(_MOBILE_CURL_CFFI_PROFILES_ALL)


def _random_mobile_user_agent() -> str:
    return random.choice(MOBILE_USER_AGENTS)


def _mobile_client_hints(ua: str) -> dict:
    if "Android" in ua:
        return {"Sec-CH-UA-Mobile": "?1", "Sec-CH-UA-Platform": '"Android"'}
    return {}


def _apply_mobile_ua(headers: dict, ua: str) -> dict:
    result = dict(headers)
    result["User-Agent"] = ua
    for key in ("Sec-CH-UA-Mobile", "Sec-CH-UA-Platform"):
        result.pop(key, None)
    result.update(_mobile_client_hints(ua))
    return result


def _context_device_kwargs(mobile=None) -> dict:
    if _use_mobile(mobile):
        return dict(random.choice(MOBILE_DEVICES))
    return {"user_agent": _random_user_agent(), "viewport": _random_viewport()}


def _looks_cloaked(resp) -> bool:
    try:
        if resp.status_code >= 400:
            return False
        return len(resp.text.strip()) < CLOAK_EMPTY_BODY_LIMIT
    except Exception:
        return False


def _build_headers(referer: str = "", user_agent: str = None,
                   mobile=None) -> dict:
    mobile = _use_mobile(mobile)
    ua = user_agent or (_random_mobile_user_agent() if mobile else _random_user_agent())
    headers = {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "cross-site",
        "Sec-Fetch-User": "?1",
    }
    if mobile:
        headers.update(_mobile_client_hints(ua))
    if referer:
        headers["Referer"] = referer
    return headers


BUTTON_KEYWORDS = re.compile(
    r'(вход|войти|регистрация|зарегистрироваться|играть|play|sign\s*in|sign\s*up|register|log\s*in|официальный\s*сайт)',
    re.IGNORECASE
)

BUTTON_LINK_PATTERN = re.compile(
    r'<a\b[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>',
    re.IGNORECASE | re.DOTALL
)


def _find_button_urls(html: str, base_url: str) -> list[str]:
    found = []
    for match in BUTTON_LINK_PATTERN.finditer(html):
        href = match.group(1).strip()
        text = re.sub(r'<[^>]+>', '', match.group(2)).strip()
        if not text or not BUTTON_KEYWORDS.search(text):
            continue
        if href.startswith('#') or href.startswith('javascript:'):
            js_match = re.search(r"(?:window\.open|location\.href)\s*[\(=]\s*['\"]([^'\"]+)", match.group(0))
            if js_match:
                href = js_match.group(1)
            else:
                continue
        if href.startswith('/'):
            parsed = urlparse(base_url)
            href = f"{parsed.scheme}://{parsed.netloc}{href}"
        elif not href.startswith('http'):
            href = base_url.rstrip('/') + '/' + href
        if href not in found:
            found.append(href)
    return found


JS_REDIRECT_PATTERNS = [
    re.compile(r'(?:window\.location|location\.href|location\.replace)\s*[\(=]\s*["\']([^"\']+)', re.IGNORECASE),
    re.compile(r'<meta[^>]+http-equiv=["\']?refresh["\']?[^>]+url=([^"\'>\s]+)', re.IGNORECASE),
    re.compile(r'window\.open\s*\(\s*["\']([^"\']+)', re.IGNORECASE),
    re.compile(r'\.redirect\s*\(\s*["\']([^"\']+)', re.IGNORECASE),
]


def _find_js_redirect(html: str, base_url: str) -> str:
    if not html:
        return ""
    for pattern in JS_REDIRECT_PATTERNS:
        match = pattern.search(html[:5000])
        if not match:
            continue
        target = match.group(1).strip()
        if target.startswith('/'):
            parsed = urlparse(base_url)
            target = f"{parsed.scheme}://{parsed.netloc}{target}"
        if target.startswith('http') and extract_domain(target) != extract_domain(base_url):
            return target
    return ""


def _follow_redirect(url: str, referer: str = "", depth: int = 0,
                     mobile=None) -> str:
    if depth > 5:
        return url
    mobile = _use_mobile(mobile)
    try:
        resp = _requests_get_retry_proxy(
            url, _build_headers(referer or ALIVE_REFERERS[0], mobile=mobile),
            ALIVE_TIMEOUT, mobile=mobile)
        final = resp.url

        if resp.status_code < 400:
            target = _find_js_redirect(resp.text, final)
            if target:
                return _follow_redirect(target, referer, depth + 1, mobile=mobile)
            if REQUESTS_DEVICE_FALLBACK and _looks_cloaked(resp):
                other_final = _follow_redirect(url, referer, depth, mobile=not mobile)
                if other_final and other_final not in (url, final):
                    return other_final

        return final
    except Exception:
        return url


MIRROR_TRACKER_QUERY_KEYS = {
    "ref", "refcode", "btag", "stag", "sub_id", "subid",
    "clickid", "click_id", "encoded_url", "tracking_link",
    "aff_id", "affid", "pid", "offer_id", "offerid", "partner",
}


def _is_junk_mirror(url: str) -> bool:
    if not url:
        return False
    if "{" in url or "%7b" in url.lower():
        try:
            parsed = urlparse(url)
            head = (parsed.netloc + parsed.path).lower()
        except Exception:
            return True
        if "{" in head or "%7b" in head or not _mirror_has_required_keyword(url):
            return True
    if has_repeated_locale(url):
        return True
    try:
        parsed = urlparse(url)
        qs = urllib.parse.parse_qs(parsed.query)
    except Exception:
        return False
    keys = {k.lower() for k in qs}
    hits = keys & MIRROR_TRACKER_QUERY_KEYS
    utm_full = {"utm_source", "utm_medium", "utm_campaign"}.issubset(keys)
    if len(hits) >= 2 or (hits and utm_full):
        return True
    path = parsed.path.lower()
    if any(p in path for p in ("/affiliate/", "/preland/", "/landingpages/")):
        return True
    return False


CLOAK_PROXY_RETRIES = 3  # сколько РАЗНЫХ прокси пробуем на прокладке, до которой
CLOAK_RETRY_BUDGET = 45  # сек. общий предел на все эти попытки вместе


def _mirror_from_cloaker(base_url: str, hop: str, referer: str, mobile: bool) -> str:
    final = _follow_redirect(hop, referer, mobile=mobile)
    if (final and extract_domain(final) != extract_domain(base_url)
            and not _is_junk_mirror(final)):
        return final
    return ""


def _retry_cloaker_other_proxies(base_url: str, referer: str, mobile: bool) -> str:
    pool = _proxy_pool()
    if len(pool) < 2:
        return ""
    deadline = time.time() + CLOAK_RETRY_BUDGET
    candidates = random.sample(pool, min(CLOAK_PROXY_RETRIES, len(pool)))
    for proxy in candidates:
        if time.time() > deadline:
            break
        try:
            with _pinned_proxy(proxy):
                resp = _requests_get_retry_proxy(
                    base_url, _build_headers(referer or ALIVE_REFERERS[0], mobile=mobile),
                    ALIVE_TIMEOUT, mobile=mobile)
                hop = _find_js_redirect(resp.text, resp.url)
                if not hop:
                    continue
                mirror = _mirror_from_cloaker(base_url, hop, referer, mobile)
                if mirror:
                    return mirror
        except Exception:
            continue
    return ""


def _extract_mirror_url(html: str, base_url: str, brand: str, referer: str = "",
                        headless: bool = True, mobile=None,
                        allow_browser: bool = True) -> str:
    mobile = _use_mobile(mobile)
    hop = _find_js_redirect(html, base_url)
    if hop:
        mirror = _mirror_from_cloaker(base_url, hop, referer, mobile)
        if mirror:
            return mirror
        mirror = _retry_cloaker_other_proxies(base_url, referer, mobile)
        if mirror:
            return mirror

    button_urls = _find_button_urls(html, base_url)
    if button_urls:
        for btn_url in button_urls[:3]:
            final = _follow_redirect(btn_url, referer, mobile=mobile)
            if (final and extract_domain(final) != extract_domain(base_url)
                    and not _is_junk_mirror(final)):
                return final

    if PLAYWRIGHT_OK and allow_browser:
        mirror = _extract_mirror_playwright(base_url, referer, headless, brand, mobile)
        if mirror:
            return mirror

    if button_urls:
        final = _follow_redirect(button_urls[0], referer, mobile=mobile)
        if not _is_junk_mirror(final):
            return final
    return ""


PLAYWRIGHT_QUEUE_WAIT = 180  # сек. сколько готовы ПРОСТОЯТЬ В ОЧЕРЕДИ к пулу браузеров


def _run_in_browser_pool(fn, *args):
    started = threading.Event()

    def wrapper():
        started.set()
        return fn(*args)

    future = _playwright_executor.submit(wrapper)
    if not started.wait(PLAYWRIGHT_QUEUE_WAIT):
        future.cancel()
        return None
    try:
        return future.result(timeout=PLAYWRIGHT_HARD_TIMEOUT)
    except Exception:
        return None


def _extract_mirror_playwright(base_url: str, referer: str = "", headless: bool = True,
                               brand: str = "", mobile=None) -> str:
    mobile = _use_mobile(mobile)

    def attempt(as_mobile: bool) -> str:
        return _run_in_browser_pool(_extract_mirror_playwright_impl,
                                    base_url, referer, headless, brand, as_mobile) or ""

    mirror = attempt(mobile)
    if not mirror and PLAYWRIGHT_DEVICE_FALLBACK:
        mirror = attempt(not mobile)
    return mirror


PLAYWRIGHT_LAUNCH_ARGS = [
    "--disable-gpu",
    "--disable-software-rasterizer",
    "--disable-dev-shm-usage",
    "--no-sandbox",
    "--num-raster-threads=1",
    "--mute-audio",
    "--disable-extensions",
    "--disable-component-update",
    "--disable-background-networking",
    "--disable-client-side-phishing-detection",
    "--disable-sync",
    "--no-first-run",
    "--no-default-browser-check",
    "--disable-blink-features=AutomationControlled",
    "--force-webrtc-ip-handling-policy=disable_non_proxied_udp",
]

BROWSERS = {
    "yandex": {
        "title": "Яндекс Браузер",
        "paths": (
            r"%LOCALAPPDATA%\Yandex\YandexBrowser\Application\browser.exe",
            r"C:\Program Files\Yandex\YandexBrowser\Application\browser.exe",
            r"C:\Program Files (x86)\Yandex\YandexBrowser\Application\browser.exe",
        ),
    },
    "brave": {
        "title": "Brave",
        "paths": (
            r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe",
            r"C:\Program Files (x86)\BraveSoftware\Brave-Browser\Application\brave.exe",
            r"%LOCALAPPDATA%\BraveSoftware\Brave-Browser\Application\brave.exe",
        ),
    },
}

BRAND_BROWSER = {
    "Pinco": "brave",
}

_browser_path_cache = {}


def _browser_exe(key: str) -> str:
    if key not in _browser_path_cache:
        found = ""
        for raw in BROWSERS.get(key, {}).get("paths", ()):
            path = os.path.expandvars(raw)
            if os.path.isfile(path):
                found = path
                break
        _browser_path_cache[key] = found
    return _browser_path_cache[key]


def _effective_brand(brand: str, url: str = "") -> str:
    if brand in BRAND_NAMES:
        return brand
    return _match_any_brand(url, strict=False) or brand


def _brand_for_browser(brand: str, url: str = "") -> str:
    if brand in BRAND_BROWSER:
        return brand
    return _effective_brand(brand, url)


def _launch_kwargs(brand: str, url: str = "") -> dict:
    key = BRAND_BROWSER.get(_brand_for_browser(brand, url))
    if not key:
        return {}
    path = _browser_exe(key)
    return {"executable_path": path} if path else {}


def _browser_title(brand: str, url: str = "") -> str:
    if not _launch_kwargs(brand, url):
        return "Chromium"
    key = BRAND_BROWSER.get(_brand_for_browser(brand, url), "")
    return BROWSERS.get(key, {}).get("title", key)


def _click_result_valid(mirror_url: str) -> bool:
    return bool(
        mirror_url
        and mirror_url != "about:blank"
        and not mirror_url.startswith("chrome")
        and not _is_junk_mirror(mirror_url)
    )


def _human_click(page, btn, timeout: int = 5000):
    try:
        box = btn.bounding_box()
        if box:
            x = box["x"] + box["width"] * random.uniform(0.3, 0.7)
            y = box["y"] + box["height"] * random.uniform(0.3, 0.7)
            page.mouse.move(x, y, steps=random.randint(8, 20))
            page.wait_for_timeout(random.randint(80, 250))
    except Exception:
        pass
    btn.click(timeout=timeout)


BUTTON_CLASS_FALLBACK_SELECTORS = [
    '[class*="registration"] a',
    '[class*="signUp"] a',
    '[class*="register"] a',
    '[class*="login"] a',
]


def _find_clickable_button(page):
    try:
        elements = page.query_selector_all("a, button")
    except Exception:
        return None
    for el in elements:
        try:
            text = el.inner_text().strip()
        except Exception:
            continue
        if text and BUTTON_KEYWORDS.search(text):
            try:
                if el.is_visible():
                    return el
            except Exception:
                continue
    try:
        for selector in BUTTON_CLASS_FALLBACK_SELECTORS:
            candidate = page.query_selector(selector)
            if candidate and candidate.is_visible():
                return candidate
    except Exception:
        pass
    return None

HARDWARE_CONCURRENCY_OPTIONS = [4, 6, 8, 12, 16]
DEVICE_MEMORY_OPTIONS = [4, 8, 16]
MOBILE_HARDWARE_CONCURRENCY_OPTIONS = [4, 6, 8]
MOBILE_DEVICE_MEMORY_OPTIONS = [4, 8]


def _build_stealth_script(mobile=None) -> str:
    if _use_mobile(mobile):
        hw_concurrency = random.choice(MOBILE_HARDWARE_CONCURRENCY_OPTIONS)
        device_memory = random.choice(MOBILE_DEVICE_MEMORY_OPTIONS)
        plugins = "[]"
        platform_patch = (
            "Object.defineProperty(navigator, 'platform', "
            "{ get: () => 'Linux armv8l' });\n"
            "Object.defineProperty(navigator, 'maxTouchPoints', { get: () => 5 });"
        )
    else:
        hw_concurrency = random.choice(HARDWARE_CONCURRENCY_OPTIONS)
        device_memory = random.choice(DEVICE_MEMORY_OPTIONS)
        plugins = "[1, 2, 3, 4, 5]"
        platform_patch = ""
    noise = random.randint(1, 255)
    return f"""
Object.defineProperty(navigator, 'webdriver', {{ get: () => undefined }});
window.chrome = window.chrome || {{ runtime: {{}} }};
Object.defineProperty(navigator, 'languages', {{ get: () => ['ru-RU', 'ru'] }});
Object.defineProperty(navigator, 'plugins', {{ get: () => {plugins} }});
Object.defineProperty(navigator, 'hardwareConcurrency', {{ get: () => {hw_concurrency} }});
Object.defineProperty(navigator, 'deviceMemory', {{ get: () => {device_memory} }});
{platform_patch}

(function() {{
    const noise = {noise};
    const origToDataURL = HTMLCanvasElement.prototype.toDataURL;
    HTMLCanvasElement.prototype.toDataURL = function(...args) {{
        try {{
            const tmp = document.createElement('canvas');
            tmp.width = this.width; tmp.height = this.height;
            const tctx = tmp.getContext('2d');
            tctx.drawImage(this, 0, 0);
            const imgData = tctx.getImageData(0, 0, tmp.width, tmp.height);
            for (let i = 0; i < imgData.data.length; i += 97) {{
                imgData.data[i] = (imgData.data[i] + noise) % 256;
            }}
            tctx.putImageData(imgData, 0, 0);
            return origToDataURL.apply(tmp, args);
        }} catch (e) {{
            return origToDataURL.apply(this, args);
        }}
    }};
    const origGetImageData = CanvasRenderingContext2D.prototype.getImageData;
    CanvasRenderingContext2D.prototype.getImageData = function(...args) {{
        const imgData = origGetImageData.apply(this, args);
        for (let i = 0; i < imgData.data.length; i += 101) {{
            imgData.data[i] = (imgData.data[i] + noise) % 256;
        }}
        return imgData;
    }};
}})();
"""


def _wait_for_stable_url(page, max_wait_ms: int = None, poll_ms: int = 400) -> str:
    max_wait_ms = PW_STABLE_URL_MS if max_wait_ms is None else max_wait_ms
    last = page.url
    stable_count = 0
    elapsed = 0
    while elapsed < max_wait_ms:
        page.wait_for_timeout(poll_ms)
        elapsed += poll_ms
        current = page.url
        if current == last:
            stable_count += 1
            if stable_count >= 2:
                break
        else:
            stable_count = 0
            last = current
    return last


def _wait_for_button(page, max_wait_ms: int = None, poll_ms: int = 250):
    max_wait_ms = PW_BUTTON_WAIT_MS if max_wait_ms is None else max_wait_ms
    elapsed = 0
    while elapsed < max_wait_ms:
        btn = _find_clickable_button(page)
        if btn:
            return btn
        try:
            page.wait_for_timeout(poll_ms)
        except Exception:
            pass
        elapsed += poll_ms
    return None


PW_GOTO_TIMEOUT_MS = 6000        # первичная навигация на сайт
PW_BUTTON_WAIT_MS = 3500         # ожидание появления кнопки Вход/Регистрация
CLICK_RESULT_WAIT_MS = 2000      # сколько всего ждём результата клика
CLICK_RESULT_POLL_MS = 200       # с каким шагом проверяем, что результат уже есть
CONSOLE_REPEAT_LIMIT = 3         # сколько раз писать в трейс одно и то же сообщение консоли
PW_AFTER_CLICK_LOAD_MS = 2500    # загрузка страницы, на которую привёл клик
PW_STABLE_URL_MS = 3000          # ожидание, пока URL перестанет меняться


def _click_and_capture(context, page, btn, log_fn=None):
    url_before = page.url
    pages_before = len(context.pages)
    try:
        _human_click(page, btn)
    except Exception as e:
        if log_fn:
            log_fn(f"Клик прерван навигацией ({e}) — проверяю, куда ушла страница.")

    target = None
    elapsed = 0
    while elapsed < CLICK_RESULT_WAIT_MS:
        try:
            if len(context.pages) > pages_before:
                target = context.pages[-1]
                break
            if page.url != url_before:
                target = page
                break
        except Exception:
            pass
        try:
            page.wait_for_timeout(CLICK_RESULT_POLL_MS)
        except Exception:
            time.sleep(CLICK_RESULT_POLL_MS / 1000)
        elapsed += CLICK_RESULT_POLL_MS

    new_page = target if (target is not None and target is not page) else None
    if new_page is not None and log_fn:
        log_fn(f"Открылась новая вкладка: {new_page.url}")
    if target is None:
        target = page
    try:
        target.wait_for_load_state("domcontentloaded", timeout=PW_AFTER_CLICK_LOAD_MS)
    except Exception:
        pass
    try:
        return _wait_for_stable_url(target), new_page
    except Exception:
        return "", new_page


PINCO_RELOAD_COUNT = 3            # сколько раз перезагружаем страницу Pinco перед поиском кнопки
PINCO_RELOAD_DELAY_RANGE = (0.4, 0.9)  # сек. пауза между перезагрузками. Ужата вместе с


def _reload_if_pinco(page, brand: str, log_fn=None, deadline: float = None):
    if brand != "Pinco":
        return
    if _find_clickable_button(page):
        return
    for i in range(PINCO_RELOAD_COUNT):
        if deadline is not None and time.time() > deadline:
            if log_fn:
                log_fn("Pinco: время попытки вышло, перезагрузки прекращены")
            return
        time.sleep(random.uniform(*PINCO_RELOAD_DELAY_RANGE))
        try:
            page.reload(wait_until="domcontentloaded", timeout=PW_GOTO_TIMEOUT_MS)
        except Exception:
            pass
        if log_fn:
            log_fn(f"Pinco: перезагрузка страницы {i + 1}/{PINCO_RELOAD_COUNT}")
        if _find_clickable_button(page):
            return


def _worth_another_proxy(last_mirror: str, brand: str, pool, tried_proxies,
                          deadline: float) -> bool:
    return bool(
        brand and last_mirror
        and not matches_brand(last_mirror, brand, strict=False)
        and pool and len(tried_proxies) < len(pool)
        and time.time() < deadline
    )


def _extract_mirror_playwright_impl(base_url: str, referer: str = "", headless: bool = True,
                                    brand: str = "", mobile=None) -> str:
    deadline = time.time() + PLAYWRIGHT_HARD_TIMEOUT - PLAYWRIGHT_DEADLINE_BUFFER
    try:
        with sync_playwright() as p:
            pool = _proxy_pool()
            tried_proxies: set = set()
            proxy_attempts = MAX_PROXY_RETRIES if pool else 1
            brand_attempts = MIRROR_CLICK_RETRIES if brand else 1
            last_mirror = ""
            for proxy_attempt in range(proxy_attempts):
                if time.time() > deadline:
                    break
                proxy_raw = _pick_proxy(exclude=tried_proxies)
                if pool:
                    tried_proxies.add(proxy_raw)
                browser = p.chromium.launch(headless=headless, args=PLAYWRIGHT_LAUNCH_ARGS,
                                            proxy=_playwright_proxy_dict(proxy_raw),
                                            **_launch_kwargs(brand, base_url))
                blocked_proxy = False
                try:
                    context = browser.new_context(
                        locale="ru-RU",
                        ignore_https_errors=True,
                        **_context_device_kwargs(mobile),
                    )
                    _apply_media_block(context, brand, base_url)
                    context.add_init_script(_build_stealth_script(mobile))
                    page = context.new_page()

                    for brand_attempt in range(brand_attempts):
                        if time.time() > deadline:
                            break
                        try:
                            nav_resp = page.goto(base_url, wait_until="domcontentloaded", timeout=PW_GOTO_TIMEOUT_MS,
                                                 referer=referer or ALIVE_REFERERS[0])
                        except Exception:
                            nav_resp = None

                        if (brand_attempt == 0 and nav_resp is not None
                                and nav_resp.status in BLOCK_STATUS_CODES
                                and pool and len(tried_proxies) < len(pool)):
                            blocked_proxy = True
                            break

                        _reload_if_pinco(page, brand, deadline=deadline)

                        btn = _wait_for_button(page)

                        if not btn:
                            return last_mirror

                        mirror_url, new_page = _click_and_capture(context, page, btn)

                        valid = _click_result_valid(mirror_url)
                        if valid:
                            last_mirror = mirror_url
                        if new_page:
                            try:
                                new_page.close()
                            except Exception:
                                pass

                        if valid and (not brand or matches_brand(mirror_url, brand, strict=False)):
                            return mirror_url

                    if not blocked_proxy:
                        if _worth_another_proxy(last_mirror, brand, pool,
                                                tried_proxies, deadline):
                            continue          # тот же URL, другой IP
                        return last_mirror
                finally:
                    try:
                        browser.close()
                    except Exception:
                        pass
            return last_mirror
    except Exception:
        return ""


def _click_with_trace(base_url: str, referer: str = "", headless: bool = True,
                      brand: str = "", mobile=None) -> tuple[str, list]:
    lines: list = []

    def log(msg):
        lines.append(msg)

    deadline = time.time() + PLAYWRIGHT_HARD_TIMEOUT - PLAYWRIGHT_DEADLINE_BUFFER

    try:
        with sync_playwright() as p:
            pool = _proxy_pool()
            tried_proxies: set = set()
            proxy_attempts = MAX_PROXY_RETRIES if pool else 1
            brand_attempts = MIRROR_CLICK_RETRIES if brand else 1
            last_mirror = ""
            for proxy_attempt in range(proxy_attempts):
                if time.time() > deadline:
                    log("Дедлайн исчерпан — прекращаю попытки.")
                    break
                proxy_raw = _pick_proxy(exclude=tried_proxies)
                if pool:
                    tried_proxies.add(proxy_raw)
                browser = p.chromium.launch(headless=headless, args=PLAYWRIGHT_LAUNCH_ARGS,
                                            proxy=_playwright_proxy_dict(proxy_raw),
                                            **_launch_kwargs(brand, base_url))
                blocked_proxy = False
                try:
                    context = browser.new_context(
                        locale="ru-RU",
                        ignore_https_errors=True,
                        **_context_device_kwargs(mobile),
                    )
                    _apply_media_block(context, brand, base_url)
                    context.add_init_script(_build_stealth_script(mobile))
                    page = context.new_page()

                    page.on("response", lambda resp: (
                        log(f"{resp.status} {resp.url}  → {resp.headers.get('location')}")
                        if 300 <= resp.status < 400 and resp.headers.get("location") else None
                    ))
                    page.on("framenavigated", lambda frame: (
                        log(f"[navigate] {frame.url}") if frame == page.main_frame else None
                    ))
                    console_seen = {}

                    def on_console(msg):
                        if msg.type not in ("error", "warning"):
                            return
                        text = msg.text
                        n = console_seen.get(text, 0) + 1
                        console_seen[text] = n
                        if n <= CONSOLE_REPEAT_LIMIT:
                            log(f"[console:{msg.type}] {text}")
                        elif n == CONSOLE_REPEAT_LIMIT + 1:
                            log(f"[console:{msg.type}] {text} "
                                f"— дальше повторы этого сообщения не пишу")

                    page.on("console", on_console)

                    for brand_attempt in range(brand_attempts):
                        if time.time() > deadline:
                            log("Дедлайн исчерпан — прекращаю попытки.")
                            break
                        if brand_attempt > 0:
                            log(f"--- Повтор клика (попытка {brand_attempt + 1}/{brand_attempts}): "
                                f"предыдущий результат не похож на домен бренда '{brand}' ---")
                        log(f"Открываю {base_url} [{_browser_title(brand, base_url)}]"
                            + (f" (прокси: {_proxy_label(proxy_raw)})" if proxy_raw else ""))
                        try:
                            nav_resp = page.goto(base_url, wait_until="domcontentloaded", timeout=PW_GOTO_TIMEOUT_MS,
                                                 referer=referer or ALIVE_REFERERS[0])
                        except Exception as e:
                            log(f"Ошибка открытия страницы: {e}")
                            nav_resp = None

                        if (brand_attempt == 0 and nav_resp is not None
                                and nav_resp.status in BLOCK_STATUS_CODES
                                and pool and len(tried_proxies) < len(pool)):
                            log(f"Ответ {nav_resp.status} — похоже на блокировку прокси "
                                f"{_proxy_label(proxy_raw)}, пробую другой.")
                            blocked_proxy = True
                            break

                        _reload_if_pinco(page, brand, log_fn=log, deadline=deadline)

                        btn = _wait_for_button(page)
                        if btn:
                            try:
                                btn_text = btn.inner_text().strip()
                            except Exception:
                                btn_text = ""
                            log(f"Кнопка найдена: текст='{btn_text}' href={btn.get_attribute('href')}")
                        else:
                            log("Кнопка не найдена (нет видимого элемента с текстом "
                                "Вход/Регистрация/Войти/Играть/... и т.п.).")
                            return last_mirror, lines

                        mirror_url, new_page = _click_and_capture(context, page, btn, log)

                        log(f"Итоговый URL: {mirror_url}")
                        valid = _click_result_valid(mirror_url)
                        if mirror_url and not valid:
                            log(f"    ...но это не похоже на зеркало "
                                f"(служебный адрес или трекинговая прокладка) — не беру")
                        if valid:
                            last_mirror = mirror_url
                        if new_page:
                            try:
                                new_page.close()
                            except Exception:
                                pass

                        if mirror_url and (not brand or matches_brand(mirror_url, brand, strict=False)):
                            return mirror_url, lines

                    if not blocked_proxy:
                        if _worth_another_proxy(last_mirror, brand, pool,
                                                tried_proxies, deadline):
                            log(f"Итог {extract_domain(last_mirror)} не похож на бренд "
                                f"{brand} — повторяю клик с другого прокси")
                            continue          # тот же URL, другой IP
                        return last_mirror, lines
                finally:
                    try:
                        browser.close()
                    except Exception:
                        pass
            return last_mirror, lines
    except Exception as e:
        log(f"Ошибка: {e}")
        return "", lines


ALIVE_JITTER_MS = (150, 600)  # случайная пауза перед проверкой — однообразный,


def quick_dns_alive(domain: str) -> bool:
    return fast_dns.resolve_alive(domain)


def check_alive(url: str, brand: str = "", headless: bool = True,
                 timeout: int = ALIVE_TIMEOUT, jitter: bool = True,
                 referers: list = None, mobile=None,
                 allow_browser: bool = True) -> dict:
    if jitter:
        time.sleep(random.uniform(*ALIVE_JITTER_MS) / 1000)
    referers = referers or ALIVE_REFERERS
    mobile = _use_mobile(mobile)
    ua = _random_user_agent()
    referer_seed = brand or extract_domain(url)
    for referer_base in referers:
        try:
            resp, referer = _requests_get_via_search(url, referer_base, referer_seed,
                                                     ua, timeout, mobile=mobile)
        except requests.exceptions.Timeout:
            return {"alive": False, "status_code": 0, "final_url": "",
                    "mirror_url": "", "referer": "", "error": "timeout", "mobile": False}
        except requests.exceptions.ConnectionError:
            return {"alive": False, "status_code": 0, "final_url": "",
                    "mirror_url": "", "referer": "", "error": "connection_error", "mobile": False}
        except Exception:
            continue
        if resp.status_code < 400:
            resp, other_referer, is_mobile = _retry_other_device_if_cloaked(
                resp, url, referer_base, referer_seed, ua, timeout, mobile)
            if other_referer:
                referer = other_referer
            mirror_url = ""
            final = resp.url
            if brand and extract_domain(final) == extract_domain(url):
                mirror_url = _extract_mirror_url(resp.text, url, brand, referer,
                                                 headless, mobile=is_mobile,
                                                 allow_browser=allow_browser)
            return {
                "alive": True, "status_code": resp.status_code,
                "final_url": final,
                "mirror_url": mirror_url or final,
                "referer": referer, "error": None, "mobile": is_mobile,
            }

    try:
        resp, last_referer = _requests_get_via_search(url, referers[0], referer_seed,
                                                      ua, timeout, mobile=mobile)
        is_mobile = mobile
        if resp.status_code < 400:
            resp, other_referer, is_mobile = _retry_other_device_if_cloaked(
                resp, url, referers[0], referer_seed, ua, timeout, mobile)
            if other_referer:
                last_referer = other_referer
        mirror_url = ""
        final = resp.url
        if brand and extract_domain(final) == extract_domain(url):
            mirror_url = _extract_mirror_url(resp.text, url, brand, last_referer,
                                             headless, mobile=is_mobile,
                                             allow_browser=allow_browser)
        return {
            "alive": resp.status_code < 500,
            "status_code": resp.status_code,
            "final_url": final,
            "mirror_url": mirror_url or final,
            "referer": last_referer, "error": None, "mobile": is_mobile,
        }
    except requests.exceptions.Timeout:
        return {"alive": False, "status_code": 0, "final_url": "", "mirror_url": "", "referer": "", "error": "timeout", "mobile": False}
    except requests.exceptions.ConnectionError:
        return {"alive": False, "status_code": 0, "final_url": "", "mirror_url": "", "referer": "", "error": "connection_error", "mobile": False}
    except Exception as e:
        return {"alive": False, "status_code": 0, "final_url": "", "mirror_url": "", "referer": "", "error": str(e), "mobile": False}


def _retry_other_device_if_cloaked(resp, url: str, referer_base: str, seed: str,
                                    ua: str, timeout: int, mobile: bool):
    if not (REQUESTS_DEVICE_FALLBACK and _looks_cloaked(resp)):
        return resp, None, mobile
    try:
        if _find_js_redirect(resp.text, resp.url):
            return resp, None, mobile
        other_resp, other_referer = _requests_get_via_search(
            url, referer_base, seed, ua, timeout, mobile=not mobile)
    except Exception:
        return resp, None, mobile
    if (other_resp.status_code < 400
            and len(other_resp.text.strip()) > len(resp.text.strip())):
        return other_resp, other_referer, not mobile
    return resp, None, mobile


def filter_results(results: list[dict], brand: str, strict: bool,
                   known_domains: set, trash_domains: set,
                   seen_domains: set, log_fn=print) -> list[dict]:
    accepted = []
    for r in results:
        url = r["url"]
        domain = r["domain"]

        if has_repeated_locale(url):
            continue

        specific = _match_any_brand(url, strict)
        if specific:
            matched = specific
        elif matches_brand(url, brand, strict=strict):
            matched = brand
        else:
            continue

        if domain in trash_domains and not is_auto_approved(url):
            log_fn(f"  [-] Мусорная база: {domain}")
            continue

        if domain in known_domains:
            continue

        if domain in seen_domains:
            continue

        seen_domains.add(domain)
        r["matched_brand"] = matched
        accepted.append(r)
        if matched != brand:
            log_fn(f"  [+] {domain} (бренд: {matched})")
        else:
            log_fn(f"  [+] {domain}")

    return accepted


def probe_single_url(url: str, brand: str = "", trace: bool = True,
                     log_fn=print) -> dict:
    if "://" not in url:
        url = "https://" + url
    brand = brand or (_match_any_brand(url, strict=False) or "")

    log_fn(f"URL: {url}")
    log_fn(f"Бренд: {brand or '(не определён)'} | Trace: {trace} | "
           f"устройство: {'телефон' if FORCE_MOBILE else 'десктоп'}")
    if is_auto_approved(url):
        log_fn("[*] Домен автоодобрен (см. AUTO_APPROVE_HOST_PATTERNS): в сборщике "
               "такая ссылка идёт в Google-таблицу сразу после парсинга, без Фаз 3 и 4. "
               "Ниже — проверка только для информации, на запись она не влияет.")

    log_fn("━━━ ФАЗА 3 (как в сборщике: allow_browser=False) ━━━")
    log_fn(f"    иду по ссылке (прокси в пуле: {len(_proxy_pool())}, "
           f"таймаут {ALIVE_TIMEOUT}с на запрос) — это может занять минуту...")
    t0 = time.time()
    info = check_alive(url, brand=brand or "test", jitter=False, allow_browser=False)
    log_fn(f"    Фаза 3 заняла {time.time() - t0:.0f}с")
    mirror = info.get("mirror_url", "")
    final = info.get("final_url", "")
    log_fn(f"    alive={info['alive']} HTTP {info['status_code']} "
           f"устройство ответа: {'ТЕЛЕФОН' if info.get('mobile') else 'десктоп'}")
    log_fn(f"    final_url:  {final}")
    log_fn(f"    mirror_url: {mirror}")
    if info.get("error"):
        log_fn(f"[!] Ошибка Фазы 3: {info['error']}")

    confirmed = _is_mirror_confirmed(mirror, final, brand)
    result = {"alive_info": info, "mirror_confirmed": confirmed,
              "clicked": "", "trace": []}
    if not info.get("alive"):
        log_fn("[!] Сайт не отвечает — Фаза 4 не запускается")
        return result

    if _auto_approved_alive(url, info):
        log_fn("[+] Домен автоодобрен и отвечает 200 — в сборщике Фаза 4 для него "
               "не запускается")
        return result

    if confirmed:
        log_fn(f"[+] Зеркало подтверждено без браузера ({extract_domain(mirror)}) "
               "— сборщик на этой строке Фазу 4 пропустит")
        return result

    if mirror and mirror != final:
        for reason, failed in (
            ("домен в базе трекеров", is_tracker_domain(mirror)),
            ("повторяющаяся локаль в URL", has_repeated_locale(mirror)),
            ("нет обязательного ключевого слова", not _mirror_has_required_keyword(mirror)),
            ("не похоже на бренд", not matches_brand(mirror, brand, strict=False)),
        ):
            if failed:
                log_fn(f"    зеркало Фазы 3 отклонено: {reason}")
    else:
        log_fn("    Фаза 3 зеркала не нашла")

    log_fn(f"━━━ ФАЗА 4 (клик в браузере: {_browser_title(brand, url)}) ━━━")
    if not PLAYWRIGHT_OK:
        log_fn("[!] Playwright не установлен")
        return result
    t0 = time.time()
    clicked, lines = _run_playwright_click(url, info.get("referer", "") or "",
                                           trace, brand=brand)
    result["clicked"], result["trace"] = clicked, lines
    for ln in lines:
        log_fn(f"    {ln}")
    log_fn(f"    Фаза 4 заняла {time.time() - t0:.0f}с")
    if clicked:
        log_fn(f"[+] Зеркало по клику: {clicked}")
    else:
        log_fn("[!] Клик зеркала не дал")
    return result


def _run_playwright_click(url: str, referer: str, trace: bool,
                          headless: bool = True, brand: str = "",
                          mobile=None) -> tuple[str, list]:
    mobile = _use_mobile(mobile)
    fn = _click_with_trace if trace else _extract_mirror_playwright_impl

    def attempt(as_mobile: bool) -> tuple[str, list]:
        res = _run_in_browser_pool(fn, url, referer, headless, brand, as_mobile)
        if trace:
            return res if res else ("", [])
        return (res or ""), []

    mirror, lines = attempt(mobile)
    if not mirror and PLAYWRIGHT_DEVICE_FALLBACK:
        other_mirror, other_lines = attempt(not mobile)
        if other_lines:
            device = "десктоп" if mobile else "телефон"
            lines = lines + [f"[device] повтор клика с другого устройства: {device}"] + other_lines
        if other_mirror:
            mirror = other_mirror
    return mirror, lines


def _urls_equivalent(a: str, b: str) -> bool:
    if not a or not b:
        return a == b
    try:
        pa, pb = urlparse(a), urlparse(b)
        return (pa.scheme.lower(), pa.netloc.lower(), pa.path.rstrip("/"),
                pa.query) == (pb.scheme.lower(), pb.netloc.lower(), pb.path.rstrip("/"),
                              pb.query)
    except Exception:
        return a.rstrip("/") == b.rstrip("/")


def save_to_excel(results: list[dict], brand: str, skip_identical_mirror: bool = False) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = os.path.join(OUTPUT_DIR, f"{brand}_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    headers = ["Домен", "URL", "Заголовок", "Сниппет", "Живой", "HTTP код",
               "Финальный URL", "URL зеркала", "Рабочий реферер", "Ошибка", "Запрос", "Глубина"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="639922", end_color="639922", fill_type="solid")

    for r in results:
        ai = r.get("alive_info", {})
        mirror_url = ai.get("mirror_url", "")
        auto = _row_auto_approved(r)
        if auto:
            if mirror_url and (is_tracker_domain(mirror_url) or has_repeated_locale(mirror_url)):
                mirror_url = ""
        else:
            if ai.get("error"):
                continue
            if mirror_url and (is_tracker_domain(mirror_url) or has_repeated_locale(mirror_url)):
                continue
            if not _mirror_has_required_keyword(mirror_url):
                continue
            if (skip_identical_mirror and ai.get("status_code") != 403
                    and _urls_equivalent(mirror_url, r["url"])):
                continue
        ws.append([
            r["domain"], r["url"], r["title"], r["snippet"],
            "—" if ai.get("not_checked") else ("Да" if ai.get("alive") else "Нет"),
            ai.get("status_code", ""), ai.get("final_url", ""),
            mirror_url, ai.get("referer", ""), ai.get("error", ""),
            r.get("query", ""), r.get("depth", 0),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(filename)
    return filename


def _row_export_payload(r: dict, skip_identical_mirror: bool = False,
                         drop_403: bool = False) -> dict | None:
    ai = r.get("alive_info", {})
    mirror_url = ai.get("mirror_url", "")
    if _row_auto_approved(r):
        if mirror_url and (is_tracker_domain(mirror_url) or has_repeated_locale(mirror_url)):
            mirror_url = ""
        return {"url": r["url"], "final_url": ai.get("final_url", ""),
                "mirror": mirror_url}
    if ai.get("error"):
        return None
    if mirror_url and (is_tracker_domain(mirror_url) or has_repeated_locale(mirror_url)):
        return None
    if not _mirror_has_required_keyword(mirror_url):
        return None
    is_403 = ai.get("status_code") == 403
    if drop_403 and is_403:
        return None
    if skip_identical_mirror and not is_403 and _urls_equivalent(mirror_url, r["url"]):
        return None
    return {"url": r["url"], "final_url": ai.get("final_url", ""), "mirror": mirror_url}


SHEET_SPOOL_FILE = os.path.join(APP_DIR, "sheet_pending.jsonl")

_sheet_spool_lock = threading.Lock()


def _spool_add(payloads: list[dict], sheet_name: str) -> list[dict]:
    recs = []
    for p in payloads:
        rec = dict(p)
        rec["_sid"] = f"{time.time():.6f}-{random.randint(0, 10 ** 9)}"
        rec["_sheet"] = sheet_name
        recs.append(rec)
    with _sheet_spool_lock:
        try:
            with open(SHEET_SPOOL_FILE, "a", encoding="utf-8") as f:
                for rec in recs:
                    f.write(json.dumps(rec, ensure_ascii=False) + "\n")
                f.flush()
                os.fsync(f.fileno())
        except Exception:
            pass   # журнал — страховка; его отказ не должен ронять отправку
    return recs


def _spool_read() -> list[dict]:
    with _sheet_spool_lock:
        if not os.path.exists(SHEET_SPOOL_FILE):
            return []
        try:
            with open(SHEET_SPOOL_FILE, "r", encoding="utf-8") as f:
                lines = f.readlines()
        except Exception:
            return []
    recs = []
    for ln in lines:
        ln = ln.strip()
        if not ln:
            continue
        try:
            recs.append(json.loads(ln))
        except json.JSONDecodeError:
            continue   # обрезанная последняя строка после жёсткого убийства процесса
    return recs


def _spool_ack(sids: set):
    if not sids:
        return
    with _sheet_spool_lock:
        if not os.path.exists(SHEET_SPOOL_FILE):
            return
        try:
            with open(SHEET_SPOOL_FILE, "r", encoding="utf-8") as f:
                lines = f.readlines()
            keep = []
            for ln in lines:
                s = ln.strip()
                if not s:
                    continue
                try:
                    if json.loads(s).get("_sid") in sids:
                        continue
                except json.JSONDecodeError:
                    continue
                keep.append(s + "\n")
            tmp = SHEET_SPOOL_FILE + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                f.writelines(keep)
                f.flush()
                os.fsync(f.fileno())
            os.replace(tmp, SHEET_SPOOL_FILE)
        except Exception:
            pass


def _post_rows_to_sheet(rows: list[dict], log_fn=print,
                        sheet_name: str = "Yandex") -> tuple[bool, int]:
    if not rows:
        return True, 0
    if not RESULTS_APPS_SCRIPT_URL:
        return False, 0
    payload = json.dumps({"rows": rows, "sheet": sheet_name}).encode("utf-8")
    last_err = ""
    for attempt in range(1, 5):
        try:
            req = urllib.request.Request(
                RESULTS_APPS_SCRIPT_URL, data=payload,
                headers={"User-Agent": "Mozilla/5.0",
                         "Content-Type": "application/json"},
                method="POST")
            with urllib.request.urlopen(req, timeout=60) as resp:
                body = resp.read().decode("utf-8", "replace")
            try:
                info = json.loads(body)
            except json.JSONDecodeError:
                last_err = "ответ не JSON (страница Google вместо ответа скрипта)"
                raise ValueError(last_err)
            if info.get("error"):
                last_err = str(info["error"])
                log_fn(f"[!] Apps Script вернул ошибку: {last_err}")
                raise ValueError(last_err)
            added = info.get("added", len(rows))
            skipped = info.get("skipped", 0)
            if added or skipped:
                log_fn(f"[→] Google-таблица: +{added} (дублей {skipped})")
            return True, added
        except Exception as e:
            last_err = str(e) or last_err
            if attempt < 4:
                time.sleep(min(2 ** attempt, 20))
    log_fn(f"[!] Ошибка записи в Google-таблицу ({len(rows)} стр.): {last_err}")
    return False, 0


class BaseCollector:

    def __init__(self, brand: str, log_fn=print, on_progress=None,
                 on_done=None, trace: bool = True, alive_timeout: int = ALIVE_TIMEOUT,
                 alive_workers: int = ALIVE_WORKERS, skip_identical_mirror: bool = False,
                 alive_jitter: bool = True, alive_referers: list = None,
                 record_local_rejects: bool = False, sheet_name: str = "Yandex"):
        self.brand = brand
        self.sheet_name = sheet_name
        self.trace = trace
        self.alive_timeout = alive_timeout
        self.skip_identical_mirror = skip_identical_mirror
        self.alive_workers = alive_workers
        self.alive_jitter = alive_jitter
        self.drop_403_from_sheet = True
        self.alive_referers = alive_referers or ALIVE_REFERERS
        self.record_local_rejects = record_local_rejects
        self._local_rejected_written = set()
        self.log = log_fn
        self.on_progress = on_progress or (lambda *a: None)
        self.on_done = on_done or (lambda *a: None)
        self.trash_domains = set()
        self.all_accepted = []
        self._stop = False

        self._pipeline_lock = threading.Lock()
        self._check_executor = None      # пул для requests-проверки живости
        self._check_futures = []         # futures всех отправленных на проверку строк
        self._dispatched_ids = set()     # id(r) уже отправленных — защита от дублей
        self._to_remove_ids = set()      # строки с мусорным зеркалом (найдено кликом)
        self._checked_n = 0
        self._alive_n = 0
        self._traces_dir = None
        self.start_time = None

        self._async_search_executor = None

        self._sheet_queue = None
        self._sheet_writer = None
        self._sheet_closing = False
        self._sheet_added = 0
        self._sheet_inflight = set()

    def stop(self):
        self._stop = True
        _stop_event.set()
        self.log("[!] Останавливаю. Проверка уже найденных ссылок завершится в фоне...")

    def _start_pipeline(self):
        _stop_event.clear()
        self.start_time = time.time()
        start_proxy_health_monitor(self.log)
        self._check_executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=self.alive_workers, thread_name_prefix="check")
        if self.trace and PLAYWRIGHT_OK:
            self._traces_dir = os.path.join(OUTPUT_DIR, "traces")
            os.makedirs(self._traces_dir, exist_ok=True)
        self._sheet_queue = queue.Queue()
        self._sheet_closing = False
        self._sheet_writer = threading.Thread(
            target=self._sheet_writer_loop, daemon=True, name="sheet-writer")
        self._sheet_writer.start()
        if not RESULTS_APPS_SCRIPT_URL:
            self.log("[!] RESULTS_APPS_SCRIPT_URL не задан — строки копятся в "
                     f"{os.path.basename(SHEET_SPOOL_FILE)} и уйдут в таблицу, "
                     f"когда URL будет задан")
        self._requeue_spool_backlog(initial=True)
        self.log(f"\n{'━' * 50}")
        self.log(f"ФАЗА 3+4: проверка идёт в фоне "
                 f"(живость ×{self.alive_workers}, клик ×{PLAYWRIGHT_WORKERS} браузеров)")
        self.log(f"{'━' * 50}")

    def _enqueue_sheet(self, r) -> bool:
        payload = _row_export_payload(r, skip_identical_mirror=self.skip_identical_mirror,
                                       drop_403=self.drop_403_from_sheet)
        if payload is None:
            if self.record_local_rejects:
                self._record_local_reject(r)
            return False
        for rec in _spool_add([payload], self.sheet_name):
            if self._sheet_queue is not None:
                self._sheet_queue.put(rec)
        return True

    def _record_local_reject(self, r):
        domain = r.get("domain", "")
        if not domain:
            return
        if is_auto_approved(r.get("url", "")) or is_auto_approved(domain):
            return
        with self._pipeline_lock:
            if domain in self._local_rejected_written:
                return
            self._local_rejected_written.add(domain)
        append_local_rejected(domain)

    def _requeue_spool_backlog(self, initial: bool = False):
        if self._sheet_queue is None:
            return
        backlog = [rec for rec in _spool_read()
                    if rec.get("_sid") and rec["_sid"] not in self._sheet_inflight]
        if not backlog:
            return
        if initial:
            self.log(f"[*] В журнале {len(backlog)} недоставленных строк "
                     f"с прошлых запусков — дописываю в Google-таблицу")
        for rec in backlog:
            self._sheet_inflight.add(rec["_sid"])
            self._sheet_queue.put(rec)

    def _deliver_batch(self, batch: list[dict]):
        by_sheet = {}
        for rec in batch:
            by_sheet.setdefault(rec.get("_sheet") or self.sheet_name, []).append(rec)
        for sheet_name, recs in by_sheet.items():
            rows = [{"url": r.get("url", ""), "final_url": r.get("final_url", ""),
                     "mirror": r.get("mirror", "")} for r in recs]
            sids = {r["_sid"] for r in recs if r.get("_sid")}
            delay = 5
            for attempt in range(1, 6):
                ok, added = _post_rows_to_sheet(rows, self.log, sheet_name=sheet_name)
                if ok:
                    _spool_ack(sids)
                    with self._pipeline_lock:
                        self._sheet_added += added
                    break
                if attempt < 5:
                    self.log(f"[…] Повтор записи в Google-таблицу через {delay}с "
                             f"({len(rows)} стр.)")
                    time.sleep(delay)
                    delay = min(delay * 2, 60)
            else:
                self.log(f"[!] {len(rows)} строк пока не ушли в Google-таблицу — "
                         f"лежат в {os.path.basename(SHEET_SPOOL_FILE)}, "
                         f"отправлю повторно (в т.ч. при следующем запуске)")
            self._sheet_inflight -= sids

    def _sheet_writer_loop(self):
        q = self._sheet_queue
        idle = 0
        while not (self._sheet_closing and q.empty()):
            try:
                try:
                    first = q.get(timeout=0.5)
                except queue.Empty:
                    idle += 1
                    if idle >= 60 and not self._sheet_closing:
                        idle = 0
                        self._requeue_spool_backlog()
                    continue
                idle = 0
                batch = [first]
                while len(batch) < 25:
                    try:
                        batch.append(q.get_nowait())
                    except queue.Empty:
                        break
                self._sheet_inflight.update(r["_sid"] for r in batch if r.get("_sid"))
                self._deliver_batch(batch)
            except Exception as e:
                self.log(f"[!] Сбой писателя Google-таблицы: {e} — продолжаю "
                         f"(строки остаются в журнале)")
                time.sleep(1)

    def _dispatch(self, r):
        if self._check_executor is None:
            return
        rid = id(r)
        with self._pipeline_lock:
            if rid in self._dispatched_ids:
                return
            self._dispatched_ids.add(rid)
            fut = self._check_executor.submit(self._check_row, r)
            self._check_futures.append(fut)

    def _row_brand(self, r) -> str:
        return r.get("matched_brand") or self.brand

    def _check_row(self, r):
        try:
            self._check_row_impl(r)
        except Exception as e:
            self.log(f"  [!] {r.get('domain', r.get('url', '?'))}: сбой проверки — {e}")
            if self.trace:
                self.log(traceback.format_exc())
            try:
                if not r.get("passed_check"):
                    r.setdefault("alive_info", {"alive": False, "status_code": 0,
                                                 "final_url": "", "mirror_url": "",
                                                 "referer": "", "error": "", "mobile": False})
                    if self._enqueue_sheet(r):
                        r["passed_check"] = True
            except Exception:
                pass

    def _check_row_impl(self, r):
        if is_auto_approved(r["url"]):
            r["alive_info"] = {"alive": False, "status_code": 0, "final_url": "",
                               "mirror_url": "", "referer": "", "error": "",
                               "mobile": False, "not_checked": True}
            with self._pipeline_lock:
                self._checked_n += 1
                i = self._checked_n
            self.log(f"  [{i}] {r['domain']}: автоодобрен — пишу в таблицу без проверок")
            if self._enqueue_sheet(r):
                r["passed_check"] = True
            return

        alive_info = check_alive(r["url"], brand=self._row_brand(r), timeout=self.alive_timeout,
                                  jitter=self.alive_jitter, referers=self.alive_referers,
                                  allow_browser=False)
        r["alive_info"] = alive_info

        with self._pipeline_lock:
            self._checked_n += 1
            i = self._checked_n
            if alive_info["alive"]:
                self._alive_n += 1
            alive_n = self._alive_n

        domain = r["domain"]
        if not alive_info["alive"]:
            err = alive_info.get("error") or f"HTTP {alive_info['status_code']}"
            self.log(f"  [{i}] {domain}: ✗ {err}")
            self._enqueue_sheet(r)   # мёртвые без ошибки идут в выгрузку так же, как в Excel
            return

        mirror = alive_info.get("mirror_url", "")
        final = alive_info.get("final_url", "")
        if mirror and mirror != final:
            self.log(f"  [{i}] {domain}: ✓ живой → зеркало: {extract_domain(mirror)}")
        else:
            self.log(f"  [{i}] {domain}: ✓ живой ({alive_info['status_code']})")
        self.on_progress(f"Проверено {i}, живых {alive_n}")

        self._finalize_row(r, alive_info, i)

    def _finalize_row(self, r, alive_info, i):
        domain = r["domain"]
        mirror = alive_info.get("mirror_url", "")
        final = alive_info.get("final_url", "")
        row_brand = self._row_brand(r)

        mirror_confirmed = _is_mirror_confirmed(mirror, final, row_brand)
        auto_alive = _auto_approved_alive(r["url"], alive_info)

        if auto_alive:
            self.log(f"  [{i}] {domain}: домен автоодобрен и отвечает 200 "
                     f"— пропускаю клик, пишу по результату Фазы 3")
        elif mirror_confirmed:
            self.log(f"  [{i}] {domain}: зеркало уже подтверждено без Playwright "
                     f"({extract_domain(mirror)}) — пропускаю клик")
        elif PLAYWRIGHT_OK:
            referer = alive_info.get("referer", "") or ""
            old_mirror = alive_info.get("mirror_url", "")
            clicked, lines = _run_playwright_click(r["url"], referer, self.trace, brand=row_brand)

            if self.trace and lines and self._traces_dir:
                try:
                    with open(os.path.join(self._traces_dir, f"{domain}.log"),
                              "w", encoding="utf-8") as f:
                        f.write("\n".join(lines))
                except Exception:
                    pass

            clicked_domain = extract_domain(clicked) if clicked else ""
            if clicked_domain and clicked_domain != extract_domain(final):
                if clicked_domain in self.trash_domains:
                    self.log(f"  [{i}] {domain}: зеркало {clicked_domain} — мусорная база, удаляю строку")
                    with self._pipeline_lock:
                        self._to_remove_ids.add(id(r))
                    return   # мусор — в таблицу не пишем
                old_is_real = (_looks_like_real_mirror(old_mirror)
                               and extract_domain(old_mirror) != extract_domain(final))
                if old_is_real and not _looks_like_real_mirror(clicked):
                    self.log(f"  [{i}] {domain}: клик дал {clicked_domain}, но это не похоже "
                             f"на зеркало — оставляю {extract_domain(old_mirror)}")
                else:
                    overridden = (old_mirror and extract_domain(old_mirror) != clicked_domain
                                  and extract_domain(old_mirror) != extract_domain(final))
                    alive_info["mirror_url"] = clicked
                    if overridden:
                        self.log(f"  [{i}] {domain}: ✓ клик заменяет неверное зеркало "
                                 f"({extract_domain(old_mirror)} → {clicked_domain})")
                    else:
                        self.log(f"  [{i}] {domain}: ✓ зеркало найдено кликом → {clicked_domain}")

        if self._enqueue_sheet(r):
            r["passed_check"] = True

    def _wait_pending_checks(self):
        while True:
            with self._pipeline_lock:
                pending = list(self._check_futures)
            if not pending:
                return
            for _ in concurrent.futures.as_completed(pending):
                pass
            with self._pipeline_lock:
                if len(self._check_futures) == len(pending):
                    return

    def _final_spool_flush(self):
        self._sheet_queue = None   # писателя уже нет — очередь больше не читают
        self._sheet_inflight = set()
        backlog = _spool_read()
        if not backlog:
            return
        self.log(f"[*] Дописываю остаток журнала в Google-таблицу: {len(backlog)} стр.")
        for i in range(0, len(backlog), 25):
            self._deliver_batch(backlog[i:i + 25])
        left = len(_spool_read())
        if left:
            self.log(f"[!] {left} строк так и не ушли в Google-таблицу — они "
                     f"сохранены в {os.path.basename(SHEET_SPOOL_FILE)} и уйдут "
                     f"автоматически при следующем запуске")

    def _drain_pipeline(self) -> int:
        if self._check_executor is None:
            return 0
        with self._pipeline_lock:
            pending = list(self._check_futures)
        remaining = sum(1 for f in pending if not f.done())
        self.log(f"\n{'━' * 50}")
        self.log(f"Дожидаюсь проверки оставшихся ссылок "
                 f"(в конвейере {len(pending)}, ещё не готово {remaining})...")
        self.log(f"{'━' * 50}")
        for _ in concurrent.futures.as_completed(pending):
            if self._stop:
                break
        self._check_executor.shutdown(wait=not self._stop)
        self._check_executor = None
        if self._async_search_executor is not None:
            self._async_search_executor.shutdown(wait=not self._stop)
            self._async_search_executor = None

        if self._sheet_writer is not None:
            if self._stop:
                writer = self._sheet_writer

                def _finalize_sheet_writer(writer=writer):
                    self._wait_pending_checks()
                    self._sheet_closing = True
                    writer.join()
                    self._final_spool_flush()

                threading.Thread(target=_finalize_sheet_writer, daemon=True,
                                  name="sheet-writer-finalize").start()
            else:
                self._sheet_closing = True
                self._sheet_writer.join()
                self._final_spool_flush()
            self._sheet_writer = None

        with self._pipeline_lock:
            remove = set(self._to_remove_ids)
        if remove:
            self.all_accepted = [r for r in self.all_accepted if id(r) not in remove]
        return len(remove)


class Collector(BaseCollector):
    def __init__(self, brand: str, strict: bool, log_fn=print,
                 on_progress=None, on_done=None, trace: bool = True,
                 search_mode: str = "sync"):
        super().__init__(brand, log_fn, on_progress, on_done, trace,
                          skip_identical_mirror=True, record_local_rejects=True,
                          sheet_name="Yandex")
        self.strict = strict
        self.search_mode = search_mode  # "sync" | "async" (отложенный режим Yandex Search API)
        self.sdk = _init_sdk()
        self.known_domains = set()
        self.seen_domains = set()

    def run(self):
        mode_label = "Отложенный (async)" if self.search_mode == "async" else "Синхронный"
        self.log("=" * 50)
        self.log(f"  Бренд: {self.brand} | Фильтр: {'строгий' if self.strict else 'лёгкий'} | "
                 f"Режим поиска: {mode_label}")
        self.log("=" * 50)
        if self.search_mode == "async":
            self.log(f"[*] Отложенный режим: ответ от Yandex Search API обычно "
                     f"занимает от нескольких секунд до пары минут; страница, не "
                     f"ответившая за {ASYNC_OPERATION_TIMEOUT}с, считается пустой "
                     f"и не блокирует остальные")

        pool = _proxy_pool()
        if pool:
            self.log(f"[*] Прокси включён: {len(pool)} шт. (ротация по IP)")
            if not test_proxy(self.log):
                self.log("[!] Внимание: ни один прокси не отвечает — проверки, скорее всего, будут падать")

        self.known_domains = load_known_domains(self.log)
        self.trash_domains = load_trash_domains(self.log)
        self.known_domains |= load_results_domains(self.log)
        self.trash_domains |= load_local_rejected(self.log)

        self._start_pipeline()

        if self.search_mode == "async":
            self._run_async_phase()
        else:
            self._run_sync_phases()

        removed_trash = self._drain_pipeline()

        errored = sum(1 for r in self.all_accepted if r.get("alive_info", {}).get("error"))
        saved_count = len(self.all_accepted) - errored
        if self.all_accepted:
            filename = save_to_excel(self.all_accepted, self.brand)
            self.log(f"\n[+] Сохранено: {filename}")

        self.log(f"\n{'═' * 50}")
        self.log(f"  ИТОГО найдено: {len(self.all_accepted)} ссылок")
        self.log(f"  Удалено (зеркало из мусорной базы): {removed_trash}")
        self.log(f"  Отсеяно ошибками (не попали в Excel): {errored}")
        self.log(f"  Записано в Excel: {saved_count}")
        self.log(f"  Добавлено в Google-таблицу: {self._sheet_added}")
        alive = sum(1 for r in self.all_accepted if r.get("alive_info", {}).get("alive"))
        self.log(f"  Живых: {alive}")
        self.log(f"{'═' * 50}")

        self.on_done(self.all_accepted)

    def _run_sync_phases(self):
        queries = BRAND_QUERIES.get(self.brand, [f"{self.brand} зеркало"])

        self.log(f"\n{'━' * 50}")
        self.log(f"ФАЗА 1: Базовый поиск ({len(queries)} запросов × {PAGES_PER_QUERY} страниц)")
        self.log(f"{'━' * 50}")

        half = max(1, len(queries) // 2)
        for qi, query in enumerate(queries):
            if self._stop:
                break
            if qi >= half and self._sheet_added > 0:
                self.log(f"[*] Ссылки уже добавлены в Google-таблицу — "
                         f"сокращаю базовые запросы вдвое "
                         f"({half}/{len(queries)})")
                break
            self.log(f"\n[{qi+1}/{len(queries)}] Запрос: {query}")
            self._search_query(query, PAGES_PER_QUERY, depth=0)

        if self.all_accepted and not self._stop:
            self._recursive_search()

    def _run_async_phase(self):
        queries = load_async_queries(self.brand, self.log)
        if not queries:
            self.log(f"[!] async_queries/{self.brand}.txt пуст или не найден — "
                     f"использую стандартные запросы бренда как запасной вариант")
            queries = BRAND_QUERIES.get(self.brand, [f"{self.brand} зеркало"])

        self.log(f"\n{'━' * 50}")
        self.log(f"ФАЗА 1 (async): поиск по {len(queries)} предопределённым запросам "
                 f"(стр. 1-{ASYNC_SEARCH_PAGE_COUNT})")
        self.log(f"{'━' * 50}")

        for qi, query in enumerate(queries):
            if self._stop:
                break
            self.log(f"\n[{qi+1}/{len(queries)}] Запрос: {query}")
            self._search_query(query, ASYNC_SEARCH_PAGE_COUNT, depth=0)

    def _search_query(self, query: str, max_pages: int, depth: int):
        if self.search_mode == "async":
            self._search_query_async(query, max_pages, depth)
        else:
            self._search_query_sync(query, max_pages, depth)

    def _search_query_sync(self, query: str, max_pages: int, depth: int):
        region_id, region_label = _pick_search_region()
        if region_label:
            self.log(f"  [регион: {region_label}]")
        for page in range(SEARCH_START_PAGE, SEARCH_START_PAGE + max_pages):
            if self._stop:
                return
            self.on_progress(f"Поиск: «{query[:30]}» стр.{page+1}")
            results = yandex_search(self.sdk, query, page, self.log, region=region_id)
            if not results:
                self.log(f"  Стр.{page+1}: нет результатов, стоп")
                break
            self._accept_page_results(results, query, depth, page)
            time.sleep(API_DELAY)

    def _search_query_async(self, query: str, max_pages: int, depth: int):
        region_id, region_label = _pick_search_region()
        if region_label:
            self.log(f"  [регион: {region_label}]")
        pages = list(range(ASYNC_SEARCH_START_PAGE,
                           ASYNC_SEARCH_START_PAGE + ASYNC_SEARCH_PAGE_COUNT))
        operations = {}
        for page in pages:
            if self._stop:
                return
            self.on_progress(f"Отправка (async): «{query[:30]}» стр.{page+1}")
            try:
                operations[page] = yandex_search_submit(self.sdk, query, page, region=region_id)
            except Exception as e:
                self.log(f"[!] Ошибка отправки отложенного запроса стр.{page+1}: {e}")
            time.sleep(ASYNC_SUBMIT_DELAY)

        if not operations:
            return
        self.log(f"  Отправлено {len(operations)} отложенных запросов, жду ответы "
                 f"(до {ASYNC_OPERATION_TIMEOUT}с на страницу, дальше — считаю пустой)...")

        futures = {
            self._async_search_executor.submit(
                yandex_search_await, op, self.log, ASYNC_POLL_INTERVAL): page
            for page, op in operations.items()
        }
        for fut in concurrent.futures.as_completed(futures):
            page = futures[fut]
            if self._stop:
                continue
            self.on_progress(f"Поиск (async): «{query[:30]}» стр.{page+1} готова")
            try:
                results = fut.result()
            except Exception as e:
                self.log(f"[!] Ошибка отложенного запроса стр.{page+1}: {e}")
                continue
            if not results:
                self.log(f"  Стр.{page+1} (async): нет результатов")
                continue
            self._accept_page_results(results, query, depth, page)

    def _accept_page_results(self, results: list[dict], query: str, depth: int, page: int):
        accepted = filter_results(
            results, self.brand, self.strict,
            self.known_domains, self.trash_domains,
            self.seen_domains, self.log)
        for r in accepted:
            r["query"] = query
            r["depth"] = depth
        self.all_accepted.extend(accepted)
        for r in accepted:
            self._dispatch(r)
        self.log(f"  Стр.{page+1}: +{len(accepted)} (всего {len(self.all_accepted)})")

    def _recursive_search(self):
        depth = 1
        self._wait_pending_checks()
        candidates = [r for r in self.all_accepted if r.get("passed_check")]
        searched_urls = set()

        while depth <= MAX_RECURSIVE_DEPTH and candidates and not self._stop:
            self.log(f"\n{'━' * 50}")
            self.log(f"ФАЗА 2.{depth}: Рекурсивный поиск (глубина {depth}, "
                     f"кандидатов: {len(candidates)})")
            self.log(f"{'━' * 50}")

            for r in candidates:
                if self._stop:
                    break
                url = r["url"]
                if url in searched_urls:
                    continue
                searched_urls.add(url)

                self.log(f"\n  Ищу по: {url}")
                self._search_query(url, RECURSIVE_PAGES, depth)

            self._wait_pending_checks()
            new_candidates = [r for r in self.all_accepted
                              if r.get("depth") == depth
                              and r.get("passed_check")
                              and r["url"] not in searched_urls]

            if not new_candidates:
                self.log(f"  Новых кандидатов нет, завершаю рекурсию")
                break

            candidates = new_candidates
            depth += 1

    def _start_pipeline(self):
        super()._start_pipeline()
        if self.search_mode == "async":
            self._async_search_executor = concurrent.futures.ThreadPoolExecutor(
                max_workers=ASYNC_SEARCH_WORKERS, thread_name_prefix="async-search")


class BruteforceCollector(BaseCollector):

    def __init__(self, pool_path: str, brand: str,
                 log_fn=print, on_progress=None, on_done=None, trace: bool = True):
        super().__init__(brand, log_fn, on_progress, on_done, trace, BRUTEFORCE_ALIVE_TIMEOUT,
                          alive_workers=BRUTEFORCE_ALIVE_WORKERS, skip_identical_mirror=True,
                          alive_jitter=False, alive_referers=ALIVE_REFERERS[:1],
                          sheet_name="Jija")
        self.pool_path = pool_path
        self.batch_size = BF_BATCH_SIZE
        self.known_domains = set()
        self._dispatch_executor = None
        self._resolved_pending = []
        self.pool_total_lines = 0
        self.pool_start_position = 0
        self.total_read = 0
        self.total_dispatched = 0
        self.dns_done_n = 0
        self.phase2_total = 0
        self.phase2_done = 0
        self.dns_phase_done = False
        self._dns_flush_stop = threading.Event()
        self._dns_flush_thread = None

    def _start_dns_flush_timer(self):
        self._dns_flush_stop.clear()

        def _loop():
            while not self._dns_flush_stop.wait(BF_DNS_FLUSH_INTERVAL_SEC):
                fast_dns.flush_dns_cache(lambda *a, **k: None)

        self._dns_flush_thread = threading.Thread(
            target=_loop, daemon=True, name="dns-cache-flush")
        self._dns_flush_thread.start()

    def _stop_dns_flush_timer(self):
        self._dns_flush_stop.set()
        self._dns_flush_thread = None

    def _start_pipeline(self):
        super()._start_pipeline()
        self._dispatch_executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=BRUTEFORCE_DISPATCH_WORKERS, thread_name_prefix="bf-dns")

    def _dispatch(self, r):
        if self._dispatch_executor is None:
            return
        rid = id(r)
        with self._pipeline_lock:
            if rid in self._dispatched_ids:
                return
            self._dispatched_ids.add(rid)
            fut = self._dispatch_executor.submit(self._prefilter_and_check, r)
            self._check_futures.append(fut)

    def _prefilter_and_check(self, r):
        if not quick_dns_alive(r["domain"]):
            r["alive_info"] = {"alive": False, "status_code": 0, "final_url": "",
                                "mirror_url": "", "referer": "", "error": "dns_not_found"}
            with self._pipeline_lock:
                self._checked_n += 1
                self.dns_done_n += 1
            self._enqueue_sheet(r)
            return
        self.log(f"  [dns] {r['domain']}: резолвится — в очередь на HTTP-проверку")
        with self._pipeline_lock:
            self.dns_done_n += 1
            self._resolved_pending.append(r)

    def _run_alive_check_phase(self):
        with self._pipeline_lock:
            pending = self._resolved_pending
            self._resolved_pending = []
        if not pending:
            return
        self.phase2_total = len(pending)
        self.phase2_done = 0
        self.log(f"\n{'━' * 50}")
        self.log(f"ФАЗА 2: HTTP-проверка + клик для {len(pending)} резолвящихся доменов "
                 f"(DNS-предфильтр уже закончился, ×{self.alive_workers} потоков)")
        self.log(f"{'━' * 50}")
        with self._pipeline_lock:
            futures = [self._check_executor.submit(self._check_row, r) for r in pending]
            self._check_futures.extend(futures)
        for _ in concurrent.futures.as_completed(futures):
            with self._pipeline_lock:
                self.phase2_done += 1

    def _drain_pipeline(self) -> int:
        result = super()._drain_pipeline()
        if self._dispatch_executor is not None:
            self._dispatch_executor.shutdown(wait=not self._stop)
            self._dispatch_executor = None
        return result

    def run(self):
        self.log("=" * 50)
        self.log(f"  Парсинг перебором | Бренд: {self.brand} | Пул: {self.pool_path}")
        self.log("=" * 50)

        pool = _proxy_pool()
        if pool:
            self.log(f"[*] Прокси включён: {len(pool)} шт. (ротация по IP)")
            if not test_proxy(self.log):
                self.log("[!] Внимание: ни один прокси не отвечает — проверки, скорее всего, будут падать")

        self.log("[*] DNS-резолвер: системный (getaddrinfo)")

        fast_dns.flush_dns_cache(lambda *a, **k: None)

        self.known_domains = load_known_domains(self.log)
        self.trash_domains = load_trash_domains(self.log)
        self.known_domains |= load_results_domains(self.log)

        self.pool_total_lines = bruteforce.count_lines(self.pool_path)
        self.pool_start_position = bruteforce.current_position(self.pool_path)

        self._start_pipeline()
        self._start_dns_flush_timer()
        self.log(f"[*] Проверка идёт непрерывно, без пауз между порциями чтения "
                 f"(порция чтения: {self.batch_size})")

        total_skipped = 0
        last_log_at = 0
        while not self._stop:
            batch = bruteforce.pop_batch(self.pool_path, self.batch_size)
            if not batch:
                self.log("[*] Пул пуст — перебирать больше нечего")
                break

            for domain in batch:
                self.total_read += 1
                if domain in self.trash_domains or domain in self.known_domains:
                    total_skipped += 1
                    continue
                r = {"domain": domain, "url": f"https://{domain}",
                     "title": "", "snippet": "", "query": "bruteforce", "depth": 0}
                self.all_accepted.append(r)
                self._dispatch(r)
                self.total_dispatched += 1

            self.on_progress(
                f"Отправлено {self.total_dispatched}, проверено {self._checked_n}, живых {self._alive_n}")
            if self.total_dispatched - last_log_at >= 500:
                last_log_at = self.total_dispatched
                self.log(f"  Отправлено на проверку: {self.total_dispatched} | "
                         f"пропущено (уже в базах): {total_skipped} | "
                         f"проверено: {self._checked_n} | живых: {self._alive_n}")

        self.log(f"  Отправлено на проверку: {self.total_dispatched} | "
                 f"пропущено (уже в базах): {total_skipped}")

        self._wait_pending_checks()
        self.dns_phase_done = True
        self._stop_dns_flush_timer()
        self._run_alive_check_phase()

        removed_trash = self._drain_pipeline()

        errored = sum(1 for r in self.all_accepted if r.get("alive_info", {}).get("error"))
        saved_count = len(self.all_accepted) - errored
        if self.all_accepted:
            filename = save_to_excel(self.all_accepted, self.brand, skip_identical_mirror=self.skip_identical_mirror)
            self.log(f"\n[+] Сохранено: {filename}")

        self.log(f"\n{'═' * 50}")
        self.log(f"  ИТОГО проверено доменов: {len(self.all_accepted)}")
        self.log(f"  Удалено (зеркало из мусорной базы): {removed_trash}")
        self.log(f"  Отсеяно ошибками (не попали в Excel): {errored}")
        self.log(f"  Записано в Excel: {saved_count}")
        self.log(f"  Добавлено в Google-таблицу: {self._sheet_added}")
        alive = sum(1 for r in self.all_accepted if r.get("alive_info", {}).get("alive"))
        self.log(f"  Живых: {alive}")
        self.log(f"{'═' * 50}")

        self.on_done(self.all_accepted)


class App:
    BG = "#F8F8F6"

    def __init__(self, root: tk.Tk):
        self.root = root
        self.collector = None
        self._worker = None

        self._bf_log_queue = queue.Queue()

        load_proxy_config()

        root.title("Сборщик зеркал v2")
        root.geometry("800x700")
        root.minsize(700, 600)
        root.configure(bg=self.BG)

        self._build_ui()
        self.root.after(BF_LOG_FLUSH_MS, self._flush_bf_log)
        self.root.after(BF_SPEED_UPDATE_MS, self._update_bf_speed)

    def _build_ui(self):
        BG = self.BG

        header = tk.Frame(self.root, bg=BG)
        header.pack(fill="x", padx=16, pady=(14, 8))
        tk.Label(header, text="Сборщик зеркал",
                 font=("Helvetica", 16, "bold"), bg=BG, fg="#1A1A1A").pack(side="left")

        ttk.Separator(self.root, orient="horizontal").pack(fill="x", padx=16)

        self._build_shared_settings(self.root)

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)

        tab_search = tk.Frame(self.notebook, bg=BG)
        tab_bruteforce = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(tab_search, text="Поиск (Yandex)")
        self.notebook.add(tab_bruteforce, text="Парсинг Перебором")

        self._build_search_tab(tab_search)
        self._build_bruteforce_tab(tab_bruteforce)

    def _build_shared_settings(self, root):
        BG = self.BG

        sf = tk.Frame(root, bg=BG)
        sf.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(sf, text="Настройки проверки (общие для обеих вкладок)",
                 font=("Helvetica", 12), bg=BG, fg="#555").pack(anchor="w")

        row3 = tk.Frame(sf, bg=BG)
        row3.pack(fill="x", pady=(6, 0))

        tk.Label(row3, text="Потоков проверки живости:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.alive_workers_var = tk.StringVar(value=str(ALIVE_WORKERS))
        tk.Spinbox(row3, textvariable=self.alive_workers_var, from_=1, to=300,
                   font=("Helvetica", 11), width=5).pack(side="left", padx=(6, 16))

        tk.Label(row3, text="Потоков Playwright:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.pw_workers_var = tk.StringVar(value=str(PLAYWRIGHT_WORKERS))
        tk.Spinbox(row3, textvariable=self.pw_workers_var, from_=1, to=30,
                   font=("Helvetica", 11), width=5).pack(side="left", padx=(6, 16))

        self.trace_var = tk.BooleanVar(value=True)
        tk.Checkbutton(row3, text="Trace (точный клик в Фазе 4)",
                       variable=self.trace_var, font=("Helvetica", 11),
                       bg=BG, fg="#333", activebackground=BG,
                       selectcolor="white").pack(side="left")

        row_probe = tk.Frame(sf, bg=BG)
        row_probe.pack(fill="x", pady=(6, 0))
        tk.Label(row_probe, text="Проверить ссылку:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.probe_url_var = tk.StringVar()
        probe_entry = tk.Entry(row_probe, textvariable=self.probe_url_var,
                               font=("Courier", 10), relief="solid", bd=1)
        probe_entry.pack(side="left", fill="x", expand=True, padx=(6, 6))
        self.probe_btn = tk.Button(row_probe, text="Проверить", font=("Helvetica", 10),
                                   relief="flat", bg="#3B6D9E", fg="white",
                                   activebackground="#2A4F73", activeforeground="white",
                                   cursor="hand2", padx=12,
                                   command=self._probe_url)
        self.probe_btn.pack(side="left")
        probe_entry.bind("<Return>", lambda _e: self._probe_url())

        self.proxy_rows = []  # [{"url", "var", "row_frame", "status_lbl"}]

        row4 = tk.Frame(sf, bg=BG)
        row4.pack(fill="x", pady=(6, 0))

        self.use_proxy_var = tk.BooleanVar(value=USE_PROXY)
        self.use_proxy_var.trace_add("write", lambda *a: self._save_proxies_now())
        tk.Checkbutton(row4, text="Прокси включён", variable=self.use_proxy_var,
                       font=("Helvetica", 11), bg=BG, fg="#333",
                       activebackground=BG, selectcolor="white").pack(side="left")

        tk.Label(row4, text="Добавить:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left", padx=(16, 4))
        self.proxy_entry = tk.Entry(row4, font=("Courier", 10), width=30,
                                    relief="solid", bd=1)
        self.proxy_entry.pack(side="left", padx=(0, 6))
        self.proxy_entry.bind("<Return>", lambda e: self._on_add_proxy())
        tk.Button(row4, text="+ Добавить", font=("Helvetica", 10),
                  relief="flat", bg="#639922", fg="white", cursor="hand2",
                  command=self._on_add_proxy).pack(side="left", padx=(0, 6))
        tk.Button(row4, text="Проверить все", font=("Helvetica", 10),
                  relief="flat", bg="#3B6D9E", fg="white", cursor="hand2",
                  command=self._on_check_all_proxies).pack(side="left")

        row4b = tk.Frame(sf, bg=BG)
        row4b.pack(fill="x", pady=(2, 0))
        tk.Label(row4b, text="ip:port | ip:port:user:pass | user:pass@host:port",
                 font=("Helvetica", 9), bg=BG, fg="#AAA").pack(side="left")

        row4c = tk.Frame(sf, bg=BG)
        row4c.pack(fill="x", pady=(4, 0))
        self.kw_btn = tk.Button(row4c, text="", font=("Helvetica", 10),
                                relief="flat", bg="#3B6D9E", fg="white",
                                cursor="hand2",
                                command=self._open_keywords_editor)
        self.kw_btn.pack(side="left")
        self._update_kw_button()

        list_wrap = tk.Frame(sf, bg=BG, relief="solid", bd=1)
        list_wrap.pack(fill="x", pady=(4, 6))
        self.proxy_canvas = tk.Canvas(list_wrap, bg="white", height=90,
                                      highlightthickness=0)
        proxy_scroll = tk.Scrollbar(list_wrap, orient="vertical",
                                    command=self.proxy_canvas.yview)
        self.proxy_list_inner = tk.Frame(self.proxy_canvas, bg="white")
        self.proxy_list_inner.bind(
            "<Configure>",
            lambda e: self.proxy_canvas.configure(
                scrollregion=self.proxy_canvas.bbox("all")))
        self.proxy_canvas.create_window((0, 0), window=self.proxy_list_inner,
                                        anchor="nw")
        self.proxy_canvas.configure(yscrollcommand=proxy_scroll.set)
        self.proxy_canvas.pack(side="left", fill="both", expand=True)
        proxy_scroll.pack(side="right", fill="y")

        def _on_proxy_mousewheel(event):
            self.proxy_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        def _bind_proxy_wheel(_event):
            self.proxy_canvas.bind_all("<MouseWheel>", _on_proxy_mousewheel)

        def _unbind_proxy_wheel(_event):
            self.proxy_canvas.unbind_all("<MouseWheel>")

        self.proxy_canvas.bind("<Enter>", _bind_proxy_wheel)
        self.proxy_canvas.bind("<Leave>", _unbind_proxy_wheel)

        for p in PROXY_LIST:
            self._add_proxy_row(p.get("url", ""), p.get("enabled", True))

        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=16, pady=(4, 0))

    def _build_search_tab(self, root):
        BG = self.BG

        sf = tk.Frame(root, bg=BG)
        sf.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(sf, text="Настройки поиска",
                 font=("Helvetica", 12), bg=BG, fg="#555").pack(anchor="w")

        row1 = tk.Frame(sf, bg=BG)
        row1.pack(fill="x", pady=(6, 0))

        tk.Label(row1, text="Бренд:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.brand_var = tk.StringVar(value=BRAND_NAMES[0])
        brand_cb = ttk.Combobox(row1, textvariable=self.brand_var,
                                values=BRAND_NAMES, state="readonly",
                                font=("Helvetica", 11), width=14)
        brand_cb.pack(side="left", padx=(6, 16))

        tk.Label(row1, text="Фильтрация:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.filter_var = tk.StringVar(value="Строгий")
        filter_cb = ttk.Combobox(row1, textvariable=self.filter_var,
                                 values=FILTER_LEVELS, state="readonly",
                                 font=("Helvetica", 11), width=10)
        filter_cb.pack(side="left", padx=(6, 16))

        tk.Label(row1, text="Режим поиска:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.search_mode_var = tk.StringVar(value="Синхронный")
        mode_cb = ttk.Combobox(row1, textvariable=self.search_mode_var,
                               values=["Синхронный", "Отложенный (async)"],
                               state="readonly", font=("Helvetica", 11), width=18)
        mode_cb.pack(side="left", padx=(6, 0))

        row_region = tk.Frame(sf, bg=BG)
        row_region.pack(fill="x", pady=(6, 0))

        tk.Label(row_region, text="Регион поиска:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self._region_label_to_id = {label: rid for rid, label in SEARCH_REGIONS}
        region_values = ["Ротация (все регионы)", "Без региона"] + \
            [label for _, label in SEARCH_REGIONS]
        self.region_var = tk.StringVar(value="Ротация (все регионы)")
        region_cb = ttk.Combobox(row_region, textvariable=self.region_var,
                                 values=region_values, state="readonly",
                                 font=("Helvetica", 11), width=22)
        region_cb.pack(side="left", padx=(6, 0))

        row2 = tk.Frame(sf, bg=BG)
        row2.pack(fill="x", pady=(6, 0))

        tk.Label(row2, text="Страниц/запрос:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.pages_var = tk.StringVar(value=str(PAGES_PER_QUERY))
        tk.Spinbox(row2, textvariable=self.pages_var, from_=1, to=50,
                   font=("Helvetica", 11), width=5).pack(side="left", padx=(6, 16))

        tk.Label(row2, text="Рекурсивных стр.:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.rec_pages_var = tk.StringVar(value=str(RECURSIVE_PAGES))
        tk.Spinbox(row2, textvariable=self.rec_pages_var, from_=1, to=50,
                   font=("Helvetica", 11), width=5).pack(side="left", padx=(6, 16))

        tk.Label(row2, text="Макс. глубина:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.depth_var = tk.StringVar(value=str(MAX_RECURSIVE_DEPTH))
        tk.Spinbox(row2, textvariable=self.depth_var, from_=1, to=10,
                   font=("Helvetica", 11), width=5).pack(side="left", padx=(6, 0))

        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=16, pady=(10, 0))
        bf = tk.Frame(root, bg=BG)
        bf.pack(fill="x", padx=16, pady=10)

        self.start_btn = tk.Button(
            bf, text="▶  Запустить сбор",
            font=("Helvetica", 13, "bold"),
            relief="flat", bg="#639922", fg="white",
            activebackground="#3B6D11", activeforeground="white",
            cursor="hand2", padx=20, pady=10,
            command=self._start)
        self.start_btn.pack(side="left")

        self.stop_btn = tk.Button(
            bf, text="■  Остановить",
            font=("Helvetica", 13, "bold"),
            relief="flat", bg="#E24B4A", fg="white",
            activebackground="#C0392B", activeforeground="white",
            cursor="hand2", padx=20, pady=10,
            command=self._stop, state="disabled")
        self.stop_btn.pack(side="left", padx=(10, 0))

        self.status_label = tk.Label(
            bf, text="Готов к работе", font=("Helvetica", 11),
            bg=BG, fg="#888")
        self.status_label.pack(side="right")

        self.progress_var = tk.StringVar(value="")
        tk.Label(root, textvariable=self.progress_var,
                 font=("Helvetica", 10), bg=BG, fg="#639922").pack(
            fill="x", padx=16)

        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=16, pady=(6, 0))
        lf = tk.Frame(root, bg=BG)
        lf.pack(fill="both", expand=True, padx=16, pady=(6, 12))
        tk.Label(lf, text="Лог", font=("Helvetica", 11),
                 bg=BG, fg="#888").pack(anchor="w")

        self.log_box = tk.Text(
            lf, height=20, font=("Courier", 10),
            relief="flat", bg="#1E1E1E", fg="#AAAAAA",
            state="disabled", padx=8, pady=6, wrap="word")
        self.log_box.pack(fill="both", expand=True, pady=(4, 0))
        self.log_box.tag_config("ok",   foreground="#98C379")
        self.log_box.tag_config("err",  foreground="#E06C75")
        self.log_box.tag_config("info", foreground="#E5C07B")

        sb = tk.Scrollbar(self.log_box, command=self.log_box.yview)
        sb.pack(side="right", fill="y")
        self.log_box.config(yscrollcommand=sb.set)

    def _build_bruteforce_tab(self, root):
        BG = self.BG
        self.bf_collector = None
        self._bf_worker = None
        self.bf_pool_path = None

        self.bf_queue = load_bf_queue()
        self._bf_queue_running = False
        self._bf_queue_stopped = False
        self._bf_queue_index = 0

        sf = tk.Frame(root, bg=BG)
        sf.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(sf, text="Маска перебора",
                 font=("Helvetica", 12), bg=BG, fg="#555").pack(anchor="w")

        row1 = tk.Frame(sf, bg=BG)
        row1.pack(fill="x", pady=(6, 0))

        tk.Label(row1, text="Префикс:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.bf_prefix_var = tk.StringVar(value="1xbet-")
        tk.Entry(row1, textvariable=self.bf_prefix_var, font=("Courier", 11),
                 width=12, relief="solid", bd=1).pack(side="left", padx=(6, 16))

        tk.Label(row1, text="Длина:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.bf_length_var = tk.StringVar(value="4")
        tk.Spinbox(row1, textvariable=self.bf_length_var, from_=1, to=6,
                   font=("Helvetica", 11), width=5).pack(side="left", padx=(6, 16))

        tk.Label(row1, text="Суффикс:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.bf_suffix_var = tk.StringVar(value=".top")
        tk.Entry(row1, textvariable=self.bf_suffix_var, font=("Courier", 11),
                 width=10, relief="solid", bd=1).pack(side="left", padx=(6, 16))

        tk.Label(row1, text="Символы:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.bf_charset_var = tk.StringVar(
            value=bruteforce.CHARSET_LABELS["letters_digits"])
        charset_cb = ttk.Combobox(
            row1, textvariable=self.bf_charset_var,
            values=list(bruteforce.CHARSET_LABELS.values()),
            state="readonly", font=("Helvetica", 11), width=20)
        charset_cb.pack(side="left", padx=(6, 0))

        row1b = tk.Frame(sf, bg=BG)
        row1b.pack(fill="x", pady=(2, 0))
        self.bf_preview_var = tk.StringVar(value="")
        tk.Label(row1b, textvariable=self.bf_preview_var, font=("Courier", 9),
                 bg=BG, fg="#AAA").pack(side="left")
        self.bf_prefix_var.trace_add("write", lambda *a: self._update_bf_preview())
        self.bf_suffix_var.trace_add("write", lambda *a: self._update_bf_preview())
        self.bf_length_var.trace_add("write", lambda *a: self._update_bf_preview())
        self.bf_charset_var.trace_add("write", lambda *a: self._update_bf_preview())

        row2 = tk.Frame(sf, bg=BG)
        row2.pack(fill="x", pady=(8, 0))

        tk.Label(row2, text="Бренд:", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.bf_brand_var = tk.StringVar(value=BRAND_NAMES[0])
        ttk.Combobox(row2, textvariable=self.bf_brand_var, values=BRAND_NAMES,
                     state="readonly", font=("Helvetica", 11), width=14
                     ).pack(side="left", padx=(6, 16))

        row2b = tk.Frame(sf, bg=BG)
        row2b.pack(fill="x", pady=(6, 0))
        tk.Label(row2b, text="Начать с позиции (строка в пуле):", font=("Helvetica", 11),
                 bg=BG, fg="#333").pack(side="left")
        self.bf_start_pos_var = tk.StringVar(value="")
        tk.Entry(row2b, textvariable=self.bf_start_pos_var, font=("Helvetica", 11),
                 width=10, relief="solid", bd=1).pack(side="left", padx=(6, 6))
        tk.Label(row2b, text="(пусто — продолжить с текущей позиции; файл пула не меняется, "
                              "можно свободно переставлять позицию туда-обратно между запусками)",
                 font=("Helvetica", 9), bg=BG, fg="#888").pack(side="left")

        row3 = tk.Frame(sf, bg=BG)
        row3.pack(fill="x", pady=(10, 0))
        tk.Button(row3, text="Подготовить пул", font=("Helvetica", 11),
                  relief="flat", bg="#3B6D9E", fg="white", cursor="hand2",
                  padx=10, pady=4, command=self._on_prepare_pool
                  ).pack(side="left")
        tk.Button(row3, text="Обновить статус", font=("Helvetica", 11),
                  relief="flat", bg="#888", fg="white", cursor="hand2",
                  padx=10, pady=4, command=self._refresh_bf_pool_status
                  ).pack(side="left", padx=(8, 0))
        self.bf_pool_status_var = tk.StringVar(
            value="Пул не подготовлен — нажми «Подготовить пул»")
        tk.Label(row3, textvariable=self.bf_pool_status_var,
                 font=("Helvetica", 10), bg=BG, fg="#555").pack(side="left", padx=(12, 0))

        self._update_bf_preview()

        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=16, pady=(10, 0))
        qf = tk.Frame(root, bg=BG)
        qf.pack(fill="x", padx=16, pady=(8, 0))

        qhead = tk.Frame(qf, bg=BG)
        qhead.pack(fill="x")
        tk.Label(qhead, text="Очередь масок (запускаются одна за другой)",
                 font=("Helvetica", 12), bg=BG, fg="#555").pack(side="left")
        tk.Button(qhead, text="+ Добавить текущую маску", font=("Helvetica", 10),
                  relief="flat", bg="#639922", fg="white", cursor="hand2",
                  command=self._on_bf_queue_add).pack(side="left", padx=(12, 0))
        tk.Button(qhead, text="Очистить", font=("Helvetica", 10),
                  relief="flat", bg="#888", fg="white", cursor="hand2",
                  command=self._on_bf_queue_clear).pack(side="left", padx=(6, 0))

        qlist_wrap = tk.Frame(qf, bg=BG, relief="solid", bd=1)
        qlist_wrap.pack(fill="x", pady=(4, 0))
        self.bf_queue_canvas = tk.Canvas(qlist_wrap, bg="white", height=90,
                                         highlightthickness=0)
        qscroll = tk.Scrollbar(qlist_wrap, orient="vertical",
                               command=self.bf_queue_canvas.yview)
        self.bf_queue_inner = tk.Frame(self.bf_queue_canvas, bg="white")
        self.bf_queue_inner.bind(
            "<Configure>",
            lambda e: self.bf_queue_canvas.configure(
                scrollregion=self.bf_queue_canvas.bbox("all")))
        self.bf_queue_canvas.create_window((0, 0), window=self.bf_queue_inner,
                                           anchor="nw")
        self.bf_queue_canvas.configure(yscrollcommand=qscroll.set)
        self.bf_queue_canvas.pack(side="left", fill="both", expand=True)
        qscroll.pack(side="right", fill="y")

        def _on_queue_wheel(event):
            self.bf_queue_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        self.bf_queue_canvas.bind(
            "<Enter>", lambda e: self.bf_queue_canvas.bind_all("<MouseWheel>", _on_queue_wheel))
        self.bf_queue_canvas.bind(
            "<Leave>", lambda e: self.bf_queue_canvas.unbind_all("<MouseWheel>"))

        self.bf_queue_status_var = tk.StringVar(value="")
        tk.Label(qf, textvariable=self.bf_queue_status_var, font=("Helvetica", 10),
                 bg=BG, fg="#3B6D9E").pack(anchor="w", pady=(2, 0))

        self._redraw_bf_queue()

        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=16, pady=(10, 0))
        bf = tk.Frame(root, bg=BG)
        bf.pack(fill="x", padx=16, pady=10)

        self.bf_start_btn = tk.Button(
            bf, text="▶  Запустить перебор",
            font=("Helvetica", 13, "bold"),
            relief="flat", bg="#639922", fg="white",
            activebackground="#3B6D11", activeforeground="white",
            cursor="hand2", padx=20, pady=10,
            command=self._start_bruteforce)
        self.bf_start_btn.pack(side="left")

        self.bf_queue_btn = tk.Button(
            bf, text="▶▶  Запустить очередь",
            font=("Helvetica", 13, "bold"),
            relief="flat", bg="#3B6D9E", fg="white",
            activebackground="#2A527A", activeforeground="white",
            cursor="hand2", padx=20, pady=10,
            command=self._start_bf_queue)
        self.bf_queue_btn.pack(side="left", padx=(10, 0))

        self.bf_stop_btn = tk.Button(
            bf, text="■  Остановить",
            font=("Helvetica", 13, "bold"),
            relief="flat", bg="#E24B4A", fg="white",
            activebackground="#C0392B", activeforeground="white",
            cursor="hand2", padx=20, pady=10,
            command=self._stop_bruteforce, state="disabled")
        self.bf_stop_btn.pack(side="left", padx=(10, 0))

        self.bf_status_label = tk.Label(
            bf, text="Готов к работе", font=("Helvetica", 11),
            bg=BG, fg="#888")
        self.bf_status_label.pack(side="right")

        self.bf_progress_var = tk.StringVar(value="")
        tk.Label(root, textvariable=self.bf_progress_var,
                 font=("Helvetica", 10), bg=BG, fg="#639922").pack(
            fill="x", padx=16)

        pf = tk.Frame(root, bg=BG)
        pf.pack(fill="x", padx=16, pady=(4, 0))
        tk.Label(pf, text="Прочитано из пула (не значит «проверено» — см. ниже):",
                 font=("Helvetica", 9), bg=BG, fg="#888").pack(anchor="w")
        self.bf_progressbar = ttk.Progressbar(
            pf, orient="horizontal", mode="determinate", maximum=100)
        self.bf_progressbar.pack(fill="x")
        self.bf_progress_pct_var = tk.StringVar(value="")
        tk.Label(pf, textvariable=self.bf_progress_pct_var,
                 font=("Helvetica", 9), bg=BG, fg="#888").pack(anchor="w")

        sf2 = tk.Frame(root, bg=BG)
        sf2.pack(fill="x", padx=16, pady=(6, 0))
        self.bf_stage_var = tk.StringVar(value="")
        tk.Label(sf2, textvariable=self.bf_stage_var,
                 font=("Helvetica", 9, "bold"), bg=BG, fg="#3B6D9E").pack(anchor="w")
        self.bf_stage_progressbar = ttk.Progressbar(
            sf2, orient="horizontal", mode="determinate", maximum=100)
        self.bf_stage_progressbar.pack(fill="x")

        self.bf_speed_var = tk.StringVar(value="")
        tk.Label(root, textvariable=self.bf_speed_var,
                 font=("Helvetica", 11, "bold"), bg=BG, fg="#1A1A1A").pack(
            fill="x", padx=16, pady=(2, 0))

        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=16, pady=(6, 0))
        lf = tk.Frame(root, bg=BG)
        lf.pack(fill="both", expand=True, padx=16, pady=(6, 12))
        tk.Label(lf, text="Лог", font=("Helvetica", 11),
                 bg=BG, fg="#888").pack(anchor="w")

        self.bf_log_box = tk.Text(
            lf, height=20, font=("Courier", 10),
            relief="flat", bg="#1E1E1E", fg="#AAAAAA",
            state="disabled", padx=8, pady=6, wrap="word")
        self.bf_log_box.pack(fill="both", expand=True, pady=(4, 0))
        self.bf_log_box.tag_config("ok",   foreground="#98C379")
        self.bf_log_box.tag_config("err",  foreground="#E06C75")
        self.bf_log_box.tag_config("info", foreground="#E5C07B")

        bf_sb = tk.Scrollbar(self.bf_log_box, command=self.bf_log_box.yview)
        bf_sb.pack(side="right", fill="y")
        self.bf_log_box.config(yscrollcommand=bf_sb.set)

    def _current_bf_task(self) -> dict:
        try:
            length = int(self.bf_length_var.get())
        except ValueError:
            length = 0
        return _clean_bf_task({
            "prefix": self.bf_prefix_var.get(), "suffix": self.bf_suffix_var.get(),
            "length": length, "charset": self._bf_charset_name(),
            "brand": self.bf_brand_var.get(),
        })

    def _apply_bf_task(self, task: dict):
        self.bf_prefix_var.set(task["prefix"])
        self.bf_suffix_var.set(task["suffix"])
        self.bf_length_var.set(str(task["length"]))
        self.bf_charset_var.set(
            bruteforce.CHARSET_LABELS.get(task["charset"], task["charset"]))
        self.bf_brand_var.set(task["brand"])

    def _redraw_bf_queue(self):
        for child in self.bf_queue_inner.winfo_children():
            child.destroy()
        if not self.bf_queue:
            tk.Label(self.bf_queue_inner,
                     text="Пусто — задай маску выше и нажми «+ Добавить текущую маску»",
                     font=("Helvetica", 10), bg="white", fg="#AAA").pack(anchor="w", padx=6, pady=6)
        for i, task in enumerate(self.bf_queue):
            row = tk.Frame(self.bf_queue_inner, bg="white")
            row.pack(fill="x", padx=4, pady=1)
            running = self._bf_queue_running and i == self._bf_queue_index
            tk.Label(row, text=f"{i + 1}.", font=("Courier", 10), bg="white",
                     fg="#3B6D9E" if running else "#888", width=3, anchor="w"
                     ).pack(side="left")
            tk.Label(row, text=bf_task_label(task), font=("Courier", 10), bg="white",
                     fg="#1A1A1A" if running else "#333", anchor="w"
                     ).pack(side="left", fill="x", expand=True)
            if running:
                tk.Label(row, text="◀ идёт", font=("Helvetica", 9), bg="white",
                         fg="#639922").pack(side="right", padx=(0, 6))
            else:
                tk.Button(row, text="✕", font=("Helvetica", 9), fg="#E24B4A",
                          bg="white", relief="flat", cursor="hand2", bd=0,
                          command=lambda t=task: self._on_bf_queue_remove(t)
                          ).pack(side="right", padx=(0, 4))

    def _save_bf_queue_now(self):
        if not save_bf_queue(self.bf_queue):
            self._log_bf(f"[!] Не смог сохранить очередь в {BF_QUEUE_FILE}")

    def _on_bf_queue_add(self):
        task = self._current_bf_task()
        if not task:
            messagebox.showerror("Ошибка", "Длина должна быть числом больше нуля")
            return
        if task in self.bf_queue:
            messagebox.showinfo("Уже в очереди", "Такая маска в очереди уже есть")
            return
        self.bf_queue.append(task)
        self._save_bf_queue_now()
        self._redraw_bf_queue()
        self._log_bf(f"[*] В очередь добавлено: {bf_task_label(task)}")

    def _on_bf_queue_remove(self, task: dict):
        if task in self.bf_queue:
            self.bf_queue.remove(task)
            self._save_bf_queue_now()
            self._redraw_bf_queue()

    def _on_bf_queue_clear(self):
        if not self.bf_queue:
            return
        if not messagebox.askyesno("Очистить очередь",
                                   f"Удалить все маски из очереди ({len(self.bf_queue)})?"):
            return
        self.bf_queue = []
        self._save_bf_queue_now()
        self._redraw_bf_queue()

    def _start_bf_queue(self):
        if not self.bf_queue:
            messagebox.showwarning("Очередь пуста",
                                   "Сначала добавь хотя бы одну маску в очередь")
            return
        self._bf_queue_running = True
        self._bf_queue_stopped = False
        self._bf_queue_index = 0
        self.bf_queue_btn.config(state="disabled")
        self._log_bf("═" * 60)
        self._log_bf(f"[*] Запускаю очередь из {len(self.bf_queue)} масок. "
                     f"Следующая стартует сама, как только у текущей закончится "
                     f"HTTP-проверка.")
        self._run_bf_queue_item()

    def _run_bf_queue_item(self):
        if self._bf_queue_stopped or self._bf_queue_index >= len(self.bf_queue):
            self._finish_bf_queue()
            return
        task = self.bf_queue[self._bf_queue_index]
        n, total = self._bf_queue_index + 1, len(self.bf_queue)
        self._apply_bf_task(task)
        self._redraw_bf_queue()
        self.bf_queue_status_var.set(f"Очередь {n}/{total}: {bf_task_label(task)} — готовлю пул...")
        self._log_bf("═" * 60)
        self._log_bf(f"[*] Очередь {n}/{total}: {bf_task_label(task)}")

        def worker():
            try:
                path, count, created = bruteforce.generate_pool_if_missing(
                    task["prefix"], task["suffix"], task["length"], task["charset"],
                    log_fn=self._log_bf)
            except Exception as e:
                self._log_bf(f"[!] Пул для {bf_task_label(task)} подготовить не удалось: {e}")
                path = None

            def _do():
                if self._bf_queue_stopped:
                    self._finish_bf_queue()
                    return
                if not path:
                    self._log_bf("[!] Пропускаю эту маску, перехожу к следующей.")
                    self._bf_queue_index += 1
                    self._run_bf_queue_item()
                    return
                self.bf_pool_path = path
                bruteforce.set_position(path, 0)
                self._log_bf("[*] Курсор пула сброшен на 0 — маска перебирается целиком.")
                self._refresh_bf_pool_status()
                self.bf_queue_status_var.set(
                    f"Очередь {n}/{total}: {bf_task_label(task)} — перебор идёт")
                self._start_bruteforce(from_queue=True)
            self.root.after(0, _do)

        threading.Thread(target=worker, daemon=True).start()

    def _finish_bf_queue(self):
        was_running = self._bf_queue_running
        self._bf_queue_running = False
        self.bf_queue_btn.config(state="normal")
        self._redraw_bf_queue()
        if not was_running:
            return
        if self._bf_queue_stopped:
            self.bf_queue_status_var.set("Очередь остановлена")
            self._log_bf("[!] Очередь остановлена.")
        else:
            self.bf_queue_status_var.set(f"Очередь пройдена целиком: {len(self.bf_queue)} масок")
            self._log_bf("═" * 60)
            self._log_bf(f"[+] Очередь пройдена целиком: {len(self.bf_queue)} масок.")

    def _bf_charset_name(self) -> str:
        label = self.bf_charset_var.get()
        for name, lbl in bruteforce.CHARSET_LABELS.items():
            if lbl == label:
                return name
        return "letters_digits"

    def _update_bf_preview(self):
        try:
            length = max(1, int(self.bf_length_var.get() or "1"))
        except ValueError:
            length = 1
        charset_name = self._bf_charset_name()
        sample = "X" * length
        total = bruteforce.combinations_count(length, charset_name)
        self.bf_preview_var.set(
            f"Пример: {self.bf_prefix_var.get()}{sample}{self.bf_suffix_var.get()}"
            f"  →  {total:,} комбинаций".replace(",", " "))

    def _on_prepare_pool(self):
        prefix = self.bf_prefix_var.get()
        suffix = self.bf_suffix_var.get()
        try:
            length = int(self.bf_length_var.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Длина должна быть числом")
            return
        charset_name = self._bf_charset_name()
        total = bruteforce.combinations_count(length, charset_name)

        path = bruteforce.pool_path(prefix, suffix, length, charset_name)
        if not os.path.exists(path) and total > 3_000_000:
            if not messagebox.askyesno(
                "Большой пул",
                f"Комбинаций: {total:,}".replace(",", " ") +
                "\nФайл может получиться очень большим и генерация — не быстрой."
                "\nПродолжить?"):
                return

        self.bf_pool_status_var.set("Готовлю пул...")

        def worker():
            p, count, created = bruteforce.generate_pool_if_missing(
                prefix, suffix, length, charset_name, log_fn=self._log_bf)
            self.bf_pool_path = p
            def _do():
                action = "создан" if created else "уже существовал"
                self.bf_pool_status_var.set(
                    f"Пул {action}: {os.path.basename(p)} | осталось: {count:,}".replace(",", " "))
            self.root.after(0, _do)

        threading.Thread(target=worker, daemon=True).start()

    def _refresh_bf_pool_status(self):
        prefix = self.bf_prefix_var.get()
        suffix = self.bf_suffix_var.get()
        try:
            length = int(self.bf_length_var.get())
        except ValueError:
            return
        charset_name = self._bf_charset_name()
        path = bruteforce.pool_path(prefix, suffix, length, charset_name)
        if not os.path.exists(path):
            self.bf_pool_status_var.set("Пул не подготовлен — нажми «Подготовить пул»")
            self.bf_pool_path = None
            return
        self.bf_pool_path = path
        position = bruteforce.current_position(path)
        remaining = bruteforce.count_lines(path) - position
        self.bf_pool_status_var.set(
            f"Пул: {os.path.basename(path)} | позиция: {position:,} | осталось: {remaining:,}"
            .replace(",", " "))

    def _log_bf(self, msg: str, end="\n"):
        ts = datetime.now().strftime("%H:%M:%S")
        tag = ""
        if "[+]" in msg or "✓" in msg:
            tag = "ok"
        elif "[!]" in msg or "✗" in msg:
            tag = "err"
        elif "ПАЧКА" in msg or "━" in msg or "═" in msg:
            tag = "info"
        self._bf_log_queue.put((ts, msg, end, tag))

    def _flush_bf_log(self):
        lines = []
        try:
            while True:
                lines.append(self._bf_log_queue.get_nowait())
        except queue.Empty:
            pass
        if lines:
            self.bf_log_box.config(state="normal")
            for ts, msg, end, tag in lines:
                self.bf_log_box.insert("end", f"{ts}  {msg}{end}", tag)
            line_count = int(self.bf_log_box.index("end-1c").split(".")[0])
            if line_count > BF_LOG_MAX_LINES:
                self.bf_log_box.delete("1.0", f"{line_count - BF_LOG_MAX_LINES}.0")
            self.bf_log_box.config(state="disabled")
        self.root.after(BF_LOG_FLUSH_MS, self._flush_bf_log)

    def _update_bf_speed(self):
        collector = getattr(self, "bf_collector", None)
        if collector is not None and collector.start_time and not collector._stop:
            elapsed = time.time() - collector.start_time
            if elapsed >= 1:
                checked_rate = collector._checked_n / elapsed
                alive_rate = collector._alive_n / elapsed
                self.bf_speed_var.set(
                    f"⚡ Средняя скорость: {checked_rate:.1f} домен/сек "
                    f"(≈{checked_rate * 3600:,.0f}/час)  •  "
                    f"живых: ≈{alive_rate * 3600:,.1f}/час".replace(",", " "))
            if collector.pool_total_lines:
                done = min(collector.pool_start_position + collector.total_read,
                           collector.pool_total_lines)
                pct = done / collector.pool_total_lines * 100
                self.bf_progressbar["value"] = pct
                self.bf_progress_pct_var.set(
                    f"{done:,} / {collector.pool_total_lines:,} "
                    f"({pct:.1f}%)".replace(",", " "))
            self._update_bf_stage(collector)
        self.root.after(BF_SPEED_UPDATE_MS, self._update_bf_speed)

    def _update_bf_stage(self, collector):
        if not collector.dns_phase_done:
            total = collector.total_dispatched
            done = collector.dns_done_n
            label = "Стадия 1/2 — DNS-проверка"
        else:
            total = collector.phase2_total
            done = collector.phase2_done
            label = "Стадия 2/2 — HTTP-проверка + клик"
        if total:
            pct = min(100, done / total * 100)
            self.bf_stage_progressbar["value"] = pct
            self.bf_stage_var.set(
                f"{label}: {done:,} / {total:,} ({pct:.1f}%)".replace(",", " "))
        else:
            self.bf_stage_progressbar["value"] = 0
            self.bf_stage_var.set(label + ": ожидание первых доменов…")

    def _set_progress_bf(self, text: str):
        self.root.after(0, self.bf_progress_var.set, text)

    def _set_busy(self, busy: bool):
        state = "disabled" if busy else "normal"
        self.start_btn.config(state=state)
        self.bf_start_btn.config(state=state)
        self.bf_queue_btn.config(
            state="disabled" if (busy or self._bf_queue_running) else "normal")

    def _start_bruteforce(self, from_queue: bool = False):
        if not self.bf_pool_path or not os.path.exists(self.bf_pool_path):
            if from_queue:
                self._log_bf("[!] Пул не найден — пропускаю маску.")
                self._bf_queue_index += 1
                self._run_bf_queue_item()
                return
            messagebox.showwarning(
                "Нет пула", "Сначала нажми «Подготовить пул»")
            return

        brand = self.bf_brand_var.get()
        start_pos_raw = "" if from_queue else self.bf_start_pos_var.get().strip()
        if start_pos_raw:
            try:
                start_pos = max(0, int(start_pos_raw))
            except ValueError:
                messagebox.showerror("Ошибка", "Позиция должна быть числом")
                return
            actual = bruteforce.set_position(self.bf_pool_path, start_pos)
            self._log_bf(f"[*] Позиция курсора переставлена на {actual:,} "
                         f"(файл пула не менялся)".replace(",", " "))
            self._refresh_bf_pool_status()

        global ALIVE_WORKERS, PLAYWRIGHT_WORKERS, _playwright_executor
        global USE_PROXY, PROXY_LIST
        ALIVE_WORKERS = int(self.alive_workers_var.get())
        PLAYWRIGHT_WORKERS = int(self.pw_workers_var.get())
        USE_PROXY = self.use_proxy_var.get()
        PROXY_LIST = self._collect_proxy_list()
        save_proxy_config(USE_PROXY, PROXY_LIST)
        _playwright_executor.shutdown(wait=False)
        _playwright_executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=PLAYWRIGHT_WORKERS, thread_name_prefix="pw")

        self._set_busy(True)
        self.bf_stop_btn.config(state="normal")
        self.bf_status_label.config(text="Работает...", fg="#639922")
        self.bf_speed_var.set("")
        self.bf_progressbar["value"] = 0
        self.bf_progress_pct_var.set("")
        self.bf_stage_progressbar["value"] = 0
        self.bf_stage_var.set("")

        self.bf_collector = BruteforceCollector(
            pool_path=self.bf_pool_path, brand=brand,
            log_fn=self._log_bf, on_progress=self._set_progress_bf,
            on_done=self._on_done_bf, trace=self.trace_var.get(),
        )
        self._bf_worker = threading.Thread(target=self.bf_collector.run, daemon=True)
        self._bf_worker.start()

    def _stop_bruteforce(self):
        if self._bf_queue_running:
            self._bf_queue_stopped = True
            self._log_bf("[!] Останавливаю очередь — следующая маска не запустится.")
        if self.bf_collector:
            self.bf_collector.stop()
            self._log_bf("[!] Остановка...")
        self.bf_stop_btn.config(state="disabled")

    def _on_done_bf(self, results):
        def _do():
            self._set_busy(False)
            self.bf_stop_btn.config(state="disabled")
            alive = sum(1 for r in results if r.get("alive_info", {}).get("alive"))
            self.bf_status_label.config(
                text=f"Готово: {len(results)} проверено, {alive} живых",
                fg="#1A1A1A")
            self.bf_progress_var.set("")
            self.bf_speed_var.set("")
            self._refresh_bf_pool_status()
            if self._bf_queue_running:
                self._bf_queue_index += 1
                self._run_bf_queue_item()
        self.root.after(0, _do)

    def _probe_url(self):
        url = self.probe_url_var.get().strip()
        if not url:
            self._log("[!] Введите ссылку для проверки")
            return
        self._save_proxies_now()
        self.probe_btn.config(state="disabled")
        trace = self.trace_var.get()

        def worker():
            self._log("═" * 60)
            try:
                probe_single_url(url, trace=trace, log_fn=self._log)
            except Exception as e:
                self._log(f"[!] Исключение при проверке: {e}")
            finally:
                self._log("═" * 60)
                self.root.after(0, lambda: self.probe_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    def _log(self, msg: str, end="\n"):
        def _do():
            self.log_box.config(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            tag = ""
            if "[+]" in msg or "✓" in msg:
                tag = "ok"
            elif "[!]" in msg or "✗" in msg:
                tag = "err"
            elif "ФАЗА" in msg or "━" in msg or "═" in msg:
                tag = "info"
            self.log_box.insert("end", f"{ts}  {msg}{end}", tag)
            self.log_box.config(state="disabled")
        self.root.after(0, _do)

    def _set_progress(self, text: str):
        self.root.after(0, self.progress_var.set, text)

    def _add_proxy_row(self, url: str, enabled: bool = True):
        url = url.strip()
        if not url:
            return
        row = tk.Frame(self.proxy_list_inner, bg="white")
        row.pack(fill="x", padx=4, pady=1)

        var = tk.BooleanVar(value=enabled)
        var.trace_add("write", lambda *a: self._save_proxies_now())
        tk.Checkbutton(row, variable=var, bg="white",
                       activebackground="white").pack(side="left")

        label_text = _proxy_label(_normalize_proxy(url)) or url
        tk.Label(row, text=label_text, font=("Courier", 10),
                 bg="white", fg="#333", anchor="w", width=28
                 ).pack(side="left", padx=(2, 6))

        status_lbl = tk.Label(row, text="", font=("Helvetica", 9),
                              bg="white", anchor="w", justify="left",
                              wraplength=420)
        status_lbl.pack(side="left", fill="x", expand=True, padx=(0, 4))

        entry = {"url": url, "var": var, "row_frame": row,
                 "status_lbl": status_lbl}
        tk.Button(row, text="✕", font=("Helvetica", 9), fg="#E24B4A",
                  bg="white", relief="flat", cursor="hand2", bd=0,
                  command=lambda: self._remove_proxy_row(entry)
                  ).pack(side="right", padx=(0, 4))

        self.proxy_rows.append(entry)

    def _remove_proxy_row(self, entry: dict):
        entry["row_frame"].destroy()
        self.proxy_rows.remove(entry)
        self._save_proxies_now()

    def _on_add_proxy(self):
        raw = self.proxy_entry.get().strip()
        if not raw:
            return
        tokens = re.split(r'[\s,;]+', raw)
        for t in tokens:
            if t:
                self._add_proxy_row(t, enabled=True)
        self.proxy_entry.delete(0, "end")
        self._save_proxies_now()

    def _save_proxies_now(self):
        save_proxy_config(self.use_proxy_var.get(), self._collect_proxy_list())

    def _on_check_all_proxies(self):
        rows = list(self.proxy_rows)
        if not rows:
            return
        for e in rows:
            e["status_lbl"].config(text="проверяю...", fg="#888")

        def worker():
            for e in rows:
                ok, info = test_single_proxy(e["url"])
                def update(e=e, ok=ok, info=info):
                    if e not in self.proxy_rows:
                        return
                    if ok:
                        e["status_lbl"].config(text=f"✓ {info}", fg="#639922")
                    else:
                        e["status_lbl"].config(text=f"✗ {info}", fg="#E24B4A")
                self.root.after(0, update)

        threading.Thread(target=worker, daemon=True).start()

    def _collect_proxy_list(self) -> list:
        return [{"url": e["url"], "enabled": e["var"].get()}
                for e in self.proxy_rows]

    def _update_kw_button(self):
        self.kw_btn.config(
            text=f"Ключевые слова зеркал ({len(REQUIRED_MIRROR_KEYWORDS)})")

    def _open_keywords_editor(self):
        if getattr(self, "_kw_win", None) is not None and self._kw_win.winfo_exists():
            self._kw_win.lift()
            self._kw_win.focus_force()
            return

        BG = self.BG
        win = tk.Toplevel(self.root)
        self._kw_win = win
        win.title("Ключевые слова зеркал")
        win.geometry("460x520")
        win.minsize(380, 380)
        win.configure(bg=BG)
        win.transient(self.root)

        tk.Label(win, text="Ключевые слова зеркал",
                 font=("Helvetica", 14, "bold"), bg=BG, fg="#1A1A1A"
                 ).pack(anchor="w", padx=16, pady=(14, 2))
        tk.Label(win,
                 text=("Ссылка попадёт в результаты, только если в итоговом URL\n"
                       "зеркала есть хотя бы одно из этих слов.\n"
                       "Одно слово в строке. Регистр не важен."),
                 font=("Helvetica", 10), bg=BG, fg="#777", justify="left"
                 ).pack(anchor="w", padx=16, pady=(0, 8))

        text_wrap = tk.Frame(win, bg=BG)
        text_wrap.pack(fill="both", expand=True, padx=16)
        txt = tk.Text(text_wrap, font=("Courier", 11), relief="solid", bd=1,
                      wrap="none", undo=True)
        scroll = tk.Scrollbar(text_wrap, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=scroll.set)
        txt.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        txt.insert("1.0", "\n".join(REQUIRED_MIRROR_KEYWORDS))
        txt.focus_set()

        hint = tk.Label(win, text="", font=("Helvetica", 10), bg=BG, fg="#777")
        hint.pack(anchor="w", padx=16, pady=(6, 0))

        btns = tk.Frame(win, bg=BG)
        btns.pack(fill="x", padx=16, pady=12)

        def _current():
            return _clean_mirror_keywords(txt.get("1.0", "end").splitlines())

        def _on_save():
            kws = _current()
            if not kws and not messagebox.askyesno(
                    "Список пуст",
                    "Список пуст — фильтр по ключевым словам будет отключён,\n"
                    "и в результаты попадут любые зеркала.\n\nСохранить?",
                    parent=win):
                return
            applied, ok = set_mirror_keywords(kws)
            self._update_kw_button()
            if not ok:
                messagebox.showerror(
                    "Не сохранено",
                    f"Слова применены к текущему запуску, но записать\n"
                    f"{MIRROR_KEYWORDS_FILE}\nне удалось — после перезапуска "
                    f"они пропадут.", parent=win)
                return
            self._log(f"[*] Ключевые слова зеркал обновлены: {len(applied)} шт.")
            win.destroy()

        def _on_reset():
            if not messagebox.askyesno(
                    "Сбросить", "Вернуть список, зашитый в программу?",
                    parent=win):
                return
            txt.delete("1.0", "end")
            txt.insert("1.0", "\n".join(DEFAULT_REQUIRED_MIRROR_KEYWORDS))
            _refresh_hint()

        def _refresh_hint(_event=None):
            hint.config(text=f"Слов: {len(_current())}")

        txt.bind("<KeyRelease>", _refresh_hint)
        _refresh_hint()

        tk.Button(btns, text="Сохранить", font=("Helvetica", 11, "bold"),
                  relief="flat", bg="#639922", fg="white", cursor="hand2",
                  padx=16, pady=6, command=_on_save).pack(side="left")
        tk.Button(btns, text="Отмена", font=("Helvetica", 11),
                  relief="flat", bg="#DDD", fg="#333", cursor="hand2",
                  padx=16, pady=6, command=win.destroy).pack(side="left",
                                                             padx=(8, 0))
        tk.Button(btns, text="Сбросить к стандартным", font=("Helvetica", 11),
                  relief="flat", bg="#DDD", fg="#333", cursor="hand2",
                  padx=16, pady=6, command=_on_reset).pack(side="right")

    def _start(self):
        brand = self.brand_var.get()
        strict = self.filter_var.get() == "Строгий"
        search_mode = "async" if self.search_mode_var.get().startswith("Отложенный") else "sync"

        global PAGES_PER_QUERY, RECURSIVE_PAGES, MAX_RECURSIVE_DEPTH
        global ALIVE_WORKERS, PLAYWRIGHT_WORKERS, _playwright_executor
        global USE_PROXY, PROXY_LIST
        global SEARCH_REGION_MODE, SEARCH_REGION_FIXED_ID
        PAGES_PER_QUERY = int(self.pages_var.get())
        RECURSIVE_PAGES = int(self.rec_pages_var.get())
        MAX_RECURSIVE_DEPTH = int(self.depth_var.get())
        ALIVE_WORKERS = int(self.alive_workers_var.get())
        PLAYWRIGHT_WORKERS = int(self.pw_workers_var.get())
        USE_PROXY = self.use_proxy_var.get()

        region_choice = self.region_var.get()
        if region_choice == "Ротация (все регионы)":
            SEARCH_REGION_MODE = "rotate"
            SEARCH_REGION_FIXED_ID = None
        elif region_choice == "Без региона":
            SEARCH_REGION_MODE = "none"
            SEARCH_REGION_FIXED_ID = None
        else:
            SEARCH_REGION_MODE = "fixed"
            SEARCH_REGION_FIXED_ID = self._region_label_to_id.get(region_choice)
        PROXY_LIST = self._collect_proxy_list()
        save_proxy_config(USE_PROXY, PROXY_LIST)
        _playwright_executor.shutdown(wait=False)
        _playwright_executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=PLAYWRIGHT_WORKERS, thread_name_prefix="pw")

        self._set_busy(True)
        self.stop_btn.config(state="normal")
        self.status_label.config(text="Работает...", fg="#639922")

        self.collector = Collector(
            brand=brand, strict=strict,
            log_fn=self._log,
            on_progress=self._set_progress,
            on_done=self._on_done,
            trace=self.trace_var.get(),
            search_mode=search_mode,
        )

        self._worker = threading.Thread(target=self.collector.run, daemon=True)
        self._worker.start()

    def _stop(self):
        if self.collector:
            self.collector.stop()
            self._log("[!] Остановка...")
        self.stop_btn.config(state="disabled")

    def _on_done(self, results):
        def _do():
            self._set_busy(False)
            self.stop_btn.config(state="disabled")
            alive = sum(1 for r in results if r.get("alive_info", {}).get("alive"))
            self.status_label.config(
                text=f"Готово: {len(results)} найдено, {alive} живых",
                fg="#1A1A1A")
            self.progress_var.set("")
        self.root.after(0, _do)


def _playwright_chromium_ready() -> bool:
    if not PLAYWRIGHT_OK:
        return False
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            browser.close()
        return True
    except Exception:
        return False


def _install_playwright_chromium(status_cb=print) -> bool:
    status_cb("Браузер для проверки зеркал не найден — скачиваю "
               "(один раз, может занять пару минут)...")
    no_window = {"creationflags": subprocess.CREATE_NO_WINDOW} if sys.platform == "win32" else {}
    try:
        from patchright._impl._driver import compute_driver_executable, get_driver_env
        node, cli = compute_driver_executable()
        proc = subprocess.run(
            [str(node), str(cli), "install", "chromium"],
            capture_output=True, text=True, timeout=600, env=get_driver_env(), **no_window)
    except Exception:
        try:
            proc = subprocess.run(
                [sys.executable, "-m", "patchright", "install", "chromium"],
                capture_output=True, text=True, timeout=600, **no_window)
        except Exception as e:
            status_cb(f"Не удалось запустить установку браузера: {e}")
            return False
    if proc.returncode == 0:
        status_cb("Браузер установлен.")
        return True
    status_cb(f"Не удалось установить браузер (код {proc.returncode}): "
              f"{(proc.stderr or proc.stdout).strip()[-500:]}")
    return False


def _ensure_playwright_browser(root: tk.Tk):
    if not PLAYWRIGHT_OK:
        return

    splash = tk.Toplevel(root)
    splash.title("Подготовка")
    splash.geometry("420x110")
    splash.resizable(False, False)
    splash.configure(bg="#F8F8F6")
    status_var = tk.StringVar(value="Проверяю браузер для проверки зеркал...")
    tk.Label(splash, textvariable=status_var, font=("Helvetica", 10),
             bg="#F8F8F6", fg="#333", wraplength=380, justify="left").pack(
        padx=16, pady=(20, 8), fill="x")
    pb = ttk.Progressbar(splash, mode="indeterminate")
    pb.pack(padx=16, fill="x")
    pb.start(12)

    done = threading.Event()

    def _worker():
        try:
            if not _playwright_chromium_ready():
                if _install_playwright_chromium(lambda t: root.after(0, status_var.set, t)):
                    root.after(0, status_var.set, "Готово.")
                else:
                    root.after(0, status_var.set,
                               "Не удалось скачать браузер автоматически — клики по "
                               "кнопкам входа/регистрации на сайтах работать не будут, "
                               "пока не установите вручную (playwright install "
                               "chromium). Остальное работает как обычно.")
                    time.sleep(3)
        finally:
            done.set()

    threading.Thread(target=_worker, daemon=True, name="playwright-setup").start()

    def _poll():
        if done.is_set():
            pb.stop()
            if splash.winfo_exists():
                splash.destroy()
        elif splash.winfo_exists():
            root.after(150, _poll)

    root.after(150, _poll)
    if splash.winfo_exists():
        root.wait_window(splash)


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    _ensure_playwright_browser(root)
    App(root)
    root.deiconify()
    root.mainloop()
