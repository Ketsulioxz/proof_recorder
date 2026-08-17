
import os
from datetime import datetime
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OK = "ok"
PROBLEM = "problem"
DEAD = "dead"

_FILLS = {
    OK:      PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE"),
    PROBLEM: PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C"),
    DEAD:    PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE"),
}
_FONTS = {
    OK:      Font(color="006100"),
    PROBLEM: Font(color="9C6500"),
    DEAD:    Font(color="9C0006"),
}
_TITLES = {OK: "Успешно", PROBLEM: "Проблема", DEAD: "Мёртвый"}

_HEADER_FILL = PatternFill("solid", start_color="639922", end_color="639922")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

RESULT_COLUMNS = [
    ("Индекс", 10),
    ("URL", 70),
]

_INDEX_NAMES = {"index", "№", "n", "номер", "id", "nn", "п/п", "no"}
_URL_NAMES = {"url", "ссылка", "link", "адрес", "сайт", "domain", "домен"}


def status_title(outcome: str) -> str:
    return _TITLES.get(outcome, outcome)


class LinkRow:

    __slots__ = ("index", "url", "sheet_row")

    def __init__(self, index, url: str, sheet_row: int):
        self.index = index
        self.url = url
        self.sheet_row = sheet_row

    def __repr__(self):
        return f"LinkRow({self.index!r}, {self.url!r})"


def _looks_like_url(val) -> bool:
    if not isinstance(val, str):
        return False
    val = val.strip().lower()
    if not val or " " in val:
        return False
    if val.startswith(("http://", "https://")):
        return True
    head = val.split("/", 1)[0]
    return "." in head and len(head.rsplit(".", 1)[-1]) >= 2


def _find_columns(ws, scan_rows: int = 10):
    for r in range(1, min(scan_rows, ws.max_row or 1) + 1):
        idx_col = url_col = None
        for c in range(1, (ws.max_column or 1) + 1):
            val = ws.cell(row=r, column=c).value
            if not isinstance(val, str):
                continue
            name = val.strip().lower()
            if idx_col is None and name in _INDEX_NAMES:
                idx_col = c
            elif url_col is None and name in _URL_NAMES:
                url_col = c
        if url_col:
            return idx_col, url_col, r + 1

    for c in range(1, (ws.max_column or 1) + 1):
        hits = sum(1 for r in range(1, min(20, ws.max_row or 1) + 1)
                   if _looks_like_url(ws.cell(row=r, column=c).value))
        if hits >= 2 or (hits == 1 and (ws.max_row or 1) <= 2):
            return (1 if c > 1 else None), c, 1
    return None, None, 1


def read_links(path: str, log_fn=None) -> list:
    log = log_fn or (lambda m: None)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    idx_col, url_col, first_row = _find_columns(ws)
    if not url_col:
        wb.close()
        raise ValueError("В файле не нашлась колонка со ссылками (ожидается 'URL')")

    rows, seen = [], set()
    for r in range(first_row, (ws.max_row or 0) + 1):
        raw = ws.cell(row=r, column=url_col).value
        if raw is None:
            continue
        url = str(raw).strip()
        if not url or not _looks_like_url(url):
            continue
        key = url.lower().rstrip("/")
        if key in seen:
            log(f"[i] Строка {r}: {url} — дубль, пропускаю")
            continue
        seen.add(key)
        index = ws.cell(row=r, column=idx_col).value if idx_col else None
        rows.append(LinkRow(index if index is not None else len(rows) + 1, url, r))
    wb.close()
    log(f"[i] Из {os.path.basename(path)} прочитано ссылок: {len(rows)}")
    return rows


def result_path_for(output_root: str, source_path: str = "") -> str:
    stamp = f"{datetime.now():%Y-%m-%d_%H-%M}"
    base = os.path.splitext(os.path.basename(source_path))[0] if source_path else ""
    name = f"Результат_{base}_{stamp}.xlsx" if base else f"Результат_{stamp}.xlsx"
    return os.path.join(output_root, name)


class ResultWorkbook:

    def __init__(self, path: str, log_fn=None):
        self.path = path
        self.log = log_fn or (lambda m: None)
        self.counts = {OK: 0, PROBLEM: 0, DEAD: 0}
        self._pending_save = False

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Успешные"
        for i, (title, width) in enumerate(RESULT_COLUMNS, start=1):
            cell = self.ws.cell(row=1, column=i, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center")
            self.ws.column_dimensions[get_column_letter(i)].width = width
        self.ws.freeze_panes = "A2"
        self.save()

    def add(self, index, url: str, outcome: str, kind: str = "", reason: str = "",
            final_url: str = "", proxy: str = "", folder: str = ""):
        # Считаем все исходы (нужно для прогресса), но в таблицу пишем ТОЛЬКО
        # успешные ссылки — колонки Индекс и URL.
        if outcome in self.counts:
            self.counts[outcome] += 1
        if outcome != OK:
            return
        row = self.ws.max_row + 1
        for i, val in enumerate((index, url), start=1):
            self.ws.cell(row=row, column=i, value=val).alignment = Alignment(vertical="top")
        self.save()

    def save(self):
        try:
            self.wb.save(self.path)
            if self._pending_save:
                self._pending_save = False
                self.log(f"[i] Файл результата снова доступен — сохранил: {self.path}")
        except PermissionError:
            if not self._pending_save:
                self._pending_save = True
                self.log(f"[!] {os.path.basename(self.path)} открыт в Excel — сохраню, "
                         f"когда закроешь. Данные не теряются.")
        except Exception as e:
            self.log(f"[!] Не удалось сохранить файл результата: {e}")


def domain_of(url: str) -> str:
    try:
        if "://" not in url:
            url = "https://" + url
        return (urlparse(url).hostname or "").removeprefix("www.").lower()
    except Exception:
        return ""
